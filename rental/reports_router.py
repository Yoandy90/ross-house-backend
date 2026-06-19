"""
Reports Router — Financial reports for Ross House Rentals admin

Endpoints:
  GET /admin/reports/rent-roll      — point-in-time snapshot of every active unit/contract
  GET /admin/reports/t12            — trailing 12 months of income & expenses by month
  GET /admin/reports/summary        — quick JSON summary (for dashboard tiles)

Each report supports `?format=json|xlsx|pdf` (default `json`).
"""
import io
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from rental.shared import get_db, auth_admin, serialize

router = APIRouter()

PAID_STATUSES = ("completed", "paid")
PENDING_STATUSES = ("pending", "late", "partial")
MONTHS_ES = [
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
]


def _fmt_money(n: Optional[float]) -> str:
    try:
        return f"${float(n or 0):,.2f}"
    except Exception:
        return f"${n}"


def _safe_dt(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            return None
    return None


def _month_window(end: datetime) -> List[Tuple[datetime, datetime, str]]:
    """Return list of (start, end, label) for the 12 months ending in `end` (inclusive)."""
    out = []
    # Anchor to first day of `end`'s month
    cur = datetime(end.year, end.month, 1)
    for _ in range(12):
        # next-month boundary
        if cur.month == 12:
            nxt = datetime(cur.year + 1, 1, 1)
        else:
            nxt = datetime(cur.year, cur.month + 1, 1)
        label = f"{MONTHS_ES[cur.month - 1]} {cur.year}"
        out.append((cur, nxt, label))
        # step back one month
        if cur.month == 1:
            cur = datetime(cur.year - 1, 12, 1)
        else:
            cur = datetime(cur.year, cur.month - 1, 1)
    return list(reversed(out))  # oldest first


# ═══════════════════════════════════════════════════════════════════════
# DATA BUILDERS
# ═══════════════════════════════════════════════════════════════════════

async def _build_rent_roll(db, as_of: datetime) -> Dict[str, Any]:
    """Build rent roll data as-of a given date."""
    rows: List[Dict[str, Any]] = []

    contracts_query = {
        "$or": [
            {"status": "active"},
            {"end_date": {"$gte": as_of}},
        ],
    }
    contracts = []
    async for c in db.rental_contracts.find(contracts_query):
        contracts.append(c)

    # Load properties
    prop_ids = list({c.get("property_id") for c in contracts if c.get("property_id")})
    properties: Dict[str, Dict[str, Any]] = {}
    if prop_ids:
        oids = []
        for pid in prop_ids:
            try:
                oids.append(ObjectId(pid))
            except Exception:
                pass
        if oids:
            async for p in db.properties.find({"_id": {"$in": oids}}):
                properties[str(p["_id"])] = p

    for c in contracts:
        start_d = _safe_dt(c.get("start_date"))
        end_d = _safe_dt(c.get("end_date"))
        rent = float(c.get("monthly_rent", 0) or 0)
        deposit = float(c.get("security_deposit", 0) or 0)
        late_fee = float(c.get("late_fee_amount", 0) or 0)
        prop = properties.get(str(c.get("property_id", "")), {}) if c.get("property_id") else {}
        address = c.get("property_address") or prop.get("address") or prop.get("street") or "—"

        days_remaining = None
        if end_d:
            days_remaining = (end_d - as_of).days

        # Outstanding balance for this contract
        outstanding = 0.0
        try:
            async for pay in db.rental_payments.find({
                "contract_id": str(c["_id"]),
                "status": {"$in": list(PENDING_STATUSES)},
            }):
                outstanding += float(pay.get("amount", 0) or 0) + float(pay.get("late_fee", 0) or 0)
        except Exception:
            pass

        rows.append({
            "contract_id": str(c["_id"]),
            "tenant_name": c.get("tenant_name", "—"),
            "tenant_email": c.get("tenant_email", ""),
            "tenant_phone": c.get("tenant_phone", ""),
            "property_address": address,
            "property_id": c.get("property_id", ""),
            "unit": c.get("unit", "") or prop.get("unit", ""),
            "monthly_rent": rent,
            "security_deposit": deposit,
            "late_fee_amount": late_fee,
            "start_date": start_d.strftime("%Y-%m-%d") if start_d else "",
            "end_date": end_d.strftime("%Y-%m-%d") if end_d else "",
            "days_remaining": days_remaining,
            "status": c.get("status", "active"),
            "outstanding_balance": round(outstanding, 2),
        })

    # Sort by property then tenant
    rows.sort(key=lambda r: (r["property_address"], r["tenant_name"]))

    totals = {
        "units": len(rows),
        "occupied": sum(1 for r in rows if r["status"] == "active"),
        "monthly_rent": round(sum(r["monthly_rent"] for r in rows), 2),
        "security_deposits": round(sum(r["security_deposit"] for r in rows), 2),
        "outstanding": round(sum(r["outstanding_balance"] for r in rows), 2),
        "annualized_rent": round(sum(r["monthly_rent"] for r in rows) * 12, 2),
    }

    return {
        "as_of": as_of.strftime("%Y-%m-%d"),
        "rows": rows,
        "totals": totals,
    }


async def _build_t12(db, end_date: datetime) -> Dict[str, Any]:
    """Build trailing-12-month income + expense report."""
    months = _month_window(end_date)
    month_labels = [m[2] for m in months]

    # Initialize per-month buckets
    income_by_month = [0.0] * 12
    late_fees_by_month = [0.0] * 12
    expenses_by_month_total = [0.0] * 12
    expenses_by_category: Dict[str, List[float]] = defaultdict(lambda: [0.0] * 12)

    def month_idx(dt: Optional[datetime]) -> Optional[int]:
        if not dt:
            return None
        for i, (s, e, _) in enumerate(months):
            if s <= dt < e:
                return i
        return None

    # ── INCOME: paid rental_payments by payment_date ──
    start_window = months[0][0]
    end_window = months[-1][1]

    async for p in db.rental_payments.find({
        "status": {"$in": list(PAID_STATUSES)},
        "payment_date": {"$gte": start_window, "$lt": end_window},
    }):
        idx = month_idx(_safe_dt(p.get("payment_date")))
        if idx is None:
            continue
        income_by_month[idx] += float(p.get("amount", 0) or 0)
        late_fees_by_month[idx] += float(p.get("late_fee", 0) or 0)

    # ── EXPENSES: property_expenses by date (fallback created_at) ──
    expense_total = 0.0
    try:
        async for e in db.property_expenses.find({
            "$or": [
                {"date": {"$gte": start_window, "$lt": end_window}},
                {"created_at": {"$gte": start_window, "$lt": end_window}},
            ]
        }):
            dt = _safe_dt(e.get("date")) or _safe_dt(e.get("created_at"))
            idx = month_idx(dt)
            if idx is None:
                continue
            amount = float(e.get("amount", 0) or 0)
            cat = (e.get("category") or "other").lower()
            expenses_by_month_total[idx] += amount
            expenses_by_category[cat][idx] += amount
            expense_total += amount
    except Exception as ex:
        logging.warning(f"T12 expenses scan failed: {ex}")

    # NOI per month
    noi_by_month = [
        round(income_by_month[i] + late_fees_by_month[i] - expenses_by_month_total[i], 2)
        for i in range(12)
    ]

    total_income = round(sum(income_by_month), 2)
    total_late_fees = round(sum(late_fees_by_month), 2)
    total_expenses = round(sum(expenses_by_month_total), 2)
    total_noi = round(total_income + total_late_fees - total_expenses, 2)

    # Sort categories by total descending
    cat_rows = []
    for cat, arr in expenses_by_category.items():
        cat_rows.append({
            "category": cat,
            "monthly": [round(v, 2) for v in arr],
            "total": round(sum(arr), 2),
        })
    cat_rows.sort(key=lambda r: r["total"], reverse=True)

    return {
        "period_start": months[0][0].strftime("%Y-%m-%d"),
        "period_end": (end_date).strftime("%Y-%m-%d"),
        "months": month_labels,
        "income_by_month": [round(v, 2) for v in income_by_month],
        "late_fees_by_month": [round(v, 2) for v in late_fees_by_month],
        "expenses_by_month": [round(v, 2) for v in expenses_by_month_total],
        "noi_by_month": noi_by_month,
        "expenses_by_category": cat_rows,
        "totals": {
            "income": total_income,
            "late_fees": total_late_fees,
            "expenses": total_expenses,
            "noi": total_noi,
        },
    }


# ═══════════════════════════════════════════════════════════════════════
# EXPORTERS — XLSX
# ═══════════════════════════════════════════════════════════════════════

def _rent_roll_xlsx(data: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Rent Roll"

    header_fill = PatternFill("solid", fgColor="F59E0B")  # amber
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="111827", size=16)
    money_fmt = '"$"#,##0.00'
    thin = Side(border_style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws["A1"] = "Ross House Rentals — Rent Roll"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")

    ws["A2"] = f"As of: {data['as_of']}"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:I2")

    headers = [
        "Propiedad", "Unidad", "Inquilino", "Email", "Teléfono",
        "Renta Mensual", "Depósito", "Inicio", "Fin",
        "Días Restantes", "Estado", "Saldo Pendiente",
    ]
    row = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    row = 5
    for r in data["rows"]:
        vals = [
            r["property_address"], r["unit"], r["tenant_name"],
            r["tenant_email"], r["tenant_phone"],
            r["monthly_rent"], r["security_deposit"],
            r["start_date"], r["end_date"],
            r["days_remaining"] if r["days_remaining"] is not None else "",
            r["status"], r["outstanding_balance"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            if col in (6, 7, 12):
                cell.number_format = money_fmt
        row += 1

    # Totals row
    t = data["totals"]
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{t['units']} unidades")
    ws.cell(row=row, column=6, value=t["monthly_rent"]).number_format = money_fmt
    ws.cell(row=row, column=7, value=t["security_deposits"]).number_format = money_fmt
    ws.cell(row=row, column=12, value=t["outstanding"]).number_format = money_fmt
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.font = Font(bold=True)
        cell.border = border

    # Annualized footer
    row += 2
    ws.cell(row=row, column=1, value="Renta Anualizada (GPR):").font = Font(bold=True)
    ws.cell(row=row, column=6, value=t["annualized_rent"]).number_format = money_fmt
    ws.cell(row=row, column=6).font = Font(bold=True, color="059669")

    # Column widths
    widths = [32, 8, 22, 28, 16, 14, 12, 12, 12, 14, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _t12_xlsx(data: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "T-12"

    header_fill = PatternFill("solid", fgColor="F59E0B")
    cat_fill = PatternFill("solid", fgColor="FEF3C7")
    title_font = Font(bold=True, size=16, color="111827")
    money_fmt = '"$"#,##0.00'
    thin = Side(border_style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    months = data["months"]
    ncols = 1 + len(months) + 1  # label + 12 months + total

    ws.cell(row=1, column=1, value="Ross House Rentals — Trailing 12 Months (T-12)").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    ws.cell(row=2, column=1,
            value=f"Período: {data['period_start']} → {data['period_end']}").font = Font(italic=True, color="6B7280")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    # Header row
    hdr_row = 4
    ws.cell(row=hdr_row, column=1, value="Concepto")
    for i, m in enumerate(months):
        ws.cell(row=hdr_row, column=2 + i, value=m)
    ws.cell(row=hdr_row, column=ncols, value="Total")
    for col in range(1, ncols + 1):
        c = ws.cell(row=hdr_row, column=col)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    def _write_row(r: int, label: str, values: List[float], bold: bool = False, fill_label: bool = False, color: Optional[str] = None):
        ws.cell(row=r, column=1, value=label)
        if bold:
            ws.cell(row=r, column=1).font = Font(bold=True, color=color or "111827")
        if fill_label:
            ws.cell(row=r, column=1).fill = cat_fill
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=round(v, 2))
            cell.number_format = money_fmt
            cell.border = border
            if bold:
                cell.font = Font(bold=True, color=color or "111827")
        tot = ws.cell(row=r, column=ncols, value=round(sum(values), 2))
        tot.number_format = money_fmt
        tot.border = border
        if bold:
            tot.font = Font(bold=True, color=color or "111827")

    row = hdr_row + 1
    _write_row(row, "Ingresos por Renta", data["income_by_month"], bold=True, color="059669")
    row += 1
    _write_row(row, "Recargos por Atraso", data["late_fees_by_month"])
    row += 1
    total_income_row = [
        data["income_by_month"][i] + data["late_fees_by_month"][i] for i in range(12)
    ]
    _write_row(row, "TOTAL INGRESOS", total_income_row, bold=True, fill_label=True, color="059669")
    row += 2

    # Expense block
    ws.cell(row=row, column=1, value="GASTOS").font = Font(bold=True, color="DC2626")
    row += 1
    for cat in data["expenses_by_category"]:
        _write_row(row, f"  {cat['category'].capitalize()}", cat["monthly"])
        row += 1
    _write_row(row, "TOTAL GASTOS", data["expenses_by_month"], bold=True, fill_label=True, color="DC2626")
    row += 2

    # NOI
    _write_row(row, "NOI (Ingreso Neto Operativo)", data["noi_by_month"], bold=True, fill_label=True, color="1D4ED8")

    # Footer totals
    row += 2
    t = data["totals"]
    ws.cell(row=row, column=1, value="Resumen Anual").font = Font(bold=True, size=13)
    row += 1
    for label, val, color in [
        ("Ingresos totales", t["income"], "059669"),
        ("Recargos totales", t["late_fees"], "059669"),
        ("Gastos totales", t["expenses"], "DC2626"),
        ("NOI total", t["noi"], "1D4ED8"),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=val)
        cell.number_format = money_fmt
        cell.font = Font(bold=True, color=color)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "B5"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# EXPORTERS — PDF
# ═══════════════════════════════════════════════════════════════════════

def _rent_roll_pdf(data: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=landscape(letter),
        rightMargin=0.4 * inch, leftMargin=0.4 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        title="Ross House Rentals — Rent Roll",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("Title", parent=styles["Heading1"],
                            fontSize=18, textColor=colors.HexColor("#111827"),
                            spaceAfter=4, alignment=TA_CENTER)
    sub = ParagraphStyle("Sub", parent=styles["Normal"],
                          fontSize=10, textColor=colors.HexColor("#6B7280"),
                          alignment=TA_CENTER, spaceAfter=14)

    elems = [
        Paragraph("Ross House Rentals — Rent Roll", title),
        Paragraph(f"As of: <b>{data['as_of']}</b>", sub),
    ]

    headers = ["Propiedad", "Inquilino", "Renta", "Depósito", "Inicio", "Fin", "Estado", "Saldo"]
    rows = [headers]
    for r in data["rows"]:
        rows.append([
            Paragraph(r["property_address"][:55], styles["BodyText"]),
            Paragraph(r["tenant_name"], styles["BodyText"]),
            _fmt_money(r["monthly_rent"]),
            _fmt_money(r["security_deposit"]),
            r["start_date"] or "—",
            r["end_date"] or "—",
            r["status"],
            _fmt_money(r["outstanding_balance"]),
        ])
    t = data["totals"]
    rows.append([
        Paragraph(f"<b>TOTALES — {t['units']} unidades</b>", styles["BodyText"]),
        "",
        Paragraph(f"<b>{_fmt_money(t['monthly_rent'])}</b>", styles["BodyText"]),
        Paragraph(f"<b>{_fmt_money(t['security_deposits'])}</b>", styles["BodyText"]),
        "", "", "",
        Paragraph(f"<b>{_fmt_money(t['outstanding'])}</b>", styles["BodyText"]),
    ])

    table = Table(rows, colWidths=[2.6 * inch, 1.6 * inch, 0.8 * inch, 0.8 * inch,
                                    0.8 * inch, 0.8 * inch, 0.8 * inch, 0.9 * inch],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF3C7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FAFAFA")]),
    ]))
    elems.append(table)
    elems.append(Spacer(1, 14))
    elems.append(Paragraph(
        f"<b>Renta Anualizada (GPR):</b> {_fmt_money(t['annualized_rent'])}",
        ParagraphStyle("Foot", parent=styles["Normal"], fontSize=11,
                       textColor=colors.HexColor("#059669"), alignment=TA_CENTER),
    ))

    doc.build(elems)
    return bio.getvalue()


def _t12_pdf(data: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=landscape(letter),
        rightMargin=0.3 * inch, leftMargin=0.3 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        title="Ross House Rentals — T-12",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("Title", parent=styles["Heading1"],
                            fontSize=18, textColor=colors.HexColor("#111827"),
                            alignment=TA_CENTER, spaceAfter=4)
    sub = ParagraphStyle("Sub", parent=styles["Normal"],
                          fontSize=9.5, textColor=colors.HexColor("#6B7280"),
                          alignment=TA_CENTER, spaceAfter=12)

    elems = [
        Paragraph("Ross House Rentals — Trailing 12 Months (T-12)", title),
        Paragraph(f"Período: <b>{data['period_start']} → {data['period_end']}</b>", sub),
    ]

    months = data["months"]
    header = ["Concepto"] + months + ["Total"]
    rows = [header]

    def _row(label, vals, bold=False):
        r = [label] + [_fmt_money(v) for v in vals] + [_fmt_money(sum(vals))]
        return r

    rows.append(_row("Ingresos por Renta", data["income_by_month"], bold=True))
    rows.append(_row("Recargos por Atraso", data["late_fees_by_month"]))
    total_in = [data["income_by_month"][i] + data["late_fees_by_month"][i] for i in range(12)]
    rows.append(_row("TOTAL INGRESOS", total_in, bold=True))

    # spacer-ish row
    rows.append([""] * len(header))
    rows.append(["GASTOS"] + [""] * (len(header) - 1))
    for cat in data["expenses_by_category"]:
        rows.append(_row(f"  {cat['category'].capitalize()}", cat["monthly"]))
    rows.append(_row("TOTAL GASTOS", data["expenses_by_month"], bold=True))
    rows.append([""] * len(header))
    rows.append(_row("NOI", data["noi_by_month"], bold=True))

    col_widths = [1.8 * inch] + [0.65 * inch] * 12 + [0.85 * inch]
    table = Table(rows, colWidths=col_widths, repeatRows=1)

    income_idx = 1
    late_idx = 2
    total_in_idx = 3
    spacer1_idx = 4
    gastos_hdr_idx = 5
    n_cats = len(data["expenses_by_category"])
    gastos_total_idx = gastos_hdr_idx + 1 + n_cats
    spacer2_idx = gastos_total_idx + 1
    noi_idx = spacer2_idx + 1

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, total_in_idx), (-1, total_in_idx), colors.HexColor("#ECFDF5")),
        ("FONTNAME", (0, total_in_idx), (-1, total_in_idx), "Helvetica-Bold"),
        ("BACKGROUND", (0, gastos_hdr_idx), (-1, gastos_hdr_idx), colors.HexColor("#FEE2E2")),
        ("FONTNAME", (0, gastos_hdr_idx), (-1, gastos_hdr_idx), "Helvetica-Bold"),
        ("BACKGROUND", (0, gastos_total_idx), (-1, gastos_total_idx), colors.HexColor("#FEE2E2")),
        ("FONTNAME", (0, gastos_total_idx), (-1, gastos_total_idx), "Helvetica-Bold"),
        ("BACKGROUND", (0, noi_idx), (-1, noi_idx), colors.HexColor("#DBEAFE")),
        ("FONTNAME", (0, noi_idx), (-1, noi_idx), "Helvetica-Bold"),
    ]
    table.setStyle(TableStyle(style))
    elems.append(table)
    elems.append(Spacer(1, 16))

    t = data["totals"]
    summary = Table([
        ["Resumen Anual", ""],
        ["Ingresos totales", _fmt_money(t["income"])],
        ["Recargos totales", _fmt_money(t["late_fees"])],
        ["Gastos totales", _fmt_money(t["expenses"])],
        ["NOI total", _fmt_money(t["noi"])],
    ], colWidths=[3 * inch, 2 * inch])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, -1), (1, -1), colors.HexColor("#1D4ED8")),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
    ]))
    elems.append(summary)

    doc.build(elems)
    return bio.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════

@router.get("/admin/reports/rent-roll")
async def admin_rent_roll(request: Request):
    """Generate a Rent Roll report.

    Query params:
      - as_of: YYYY-MM-DD (default: today)
      - format: json (default) | xlsx | pdf
    """
    await auth_admin(request)
    qp = request.query_params
    fmt = (qp.get("format") or "json").lower()
    as_of_str = qp.get("as_of")
    if as_of_str:
        try:
            as_of = datetime.strptime(as_of_str, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="as_of inválido (YYYY-MM-DD)")
    else:
        as_of = datetime.utcnow()

    db = get_db()
    data = await _build_rent_roll(db, as_of)

    fname_date = as_of.strftime("%Y-%m-%d")
    if fmt == "xlsx":
        try:
            content = _rent_roll_xlsx(data)
        except Exception as e:
            logging.exception("rent-roll XLSX failed")
            raise HTTPException(status_code=500, detail=f"XLSX error: {e}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rent-roll_{fname_date}.xlsx"},
        )
    if fmt == "pdf":
        try:
            content = _rent_roll_pdf(data)
        except Exception as e:
            logging.exception("rent-roll PDF failed")
            raise HTTPException(status_code=500, detail=f"PDF error: {e}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=rent-roll_{fname_date}.pdf"},
        )

    return {"success": True, **data}


@router.get("/admin/reports/t12")
async def admin_t12(request: Request):
    """Generate a Trailing-12-Months income/expense report.

    Query params:
      - end_date: YYYY-MM-DD (default: today). Defines the most recent month of the window.
      - format: json (default) | xlsx | pdf
    """
    await auth_admin(request)
    qp = request.query_params
    fmt = (qp.get("format") or "json").lower()
    end_date_str = qp.get("end_date")
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="end_date inválido (YYYY-MM-DD)")
    else:
        end_date = datetime.utcnow()

    db = get_db()
    data = await _build_t12(db, end_date)

    fname_date = end_date.strftime("%Y-%m-%d")
    if fmt == "xlsx":
        try:
            content = _t12_xlsx(data)
        except Exception as e:
            logging.exception("T12 XLSX failed")
            raise HTTPException(status_code=500, detail=f"XLSX error: {e}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=t12_{fname_date}.xlsx"},
        )
    if fmt == "pdf":
        try:
            content = _t12_pdf(data)
        except Exception as e:
            logging.exception("T12 PDF failed")
            raise HTTPException(status_code=500, detail=f"PDF error: {e}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=t12_{fname_date}.pdf"},
        )

    return {"success": True, **data}


@router.get("/admin/reports/summary")
async def admin_reports_summary(request: Request):
    """Quick KPI summary for the Reports landing page (no auth-heavy work)."""
    await auth_admin(request)
    db = get_db()
    now = datetime.utcnow()

    rent_roll = await _build_rent_roll(db, now)
    t12 = await _build_t12(db, now)

    return {
        "success": True,
        "rent_roll": {
            "units": rent_roll["totals"]["units"],
            "monthly_rent": rent_roll["totals"]["monthly_rent"],
            "annualized_rent": rent_roll["totals"]["annualized_rent"],
            "outstanding": rent_roll["totals"]["outstanding"],
        },
        "t12": {
            "income": t12["totals"]["income"],
            "expenses": t12["totals"]["expenses"],
            "noi": t12["totals"]["noi"],
            "period_start": t12["period_start"],
            "period_end": t12["period_end"],
        },
    }
