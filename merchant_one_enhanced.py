"""
Merchant One Enhanced Features
AI CSV Parser, Subscription Management, Reports
"""

import os
import logging
import csv
import json
import httpx
from io import StringIO
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple
from pydantic import BaseModel
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class SmartParseRequest(BaseModel):
    """Request for AI-powered CSV parsing"""
    csvData: str
    hasHeader: bool = True
    # Optional: provide column hints
    columnHints: Optional[Dict[str, str]] = None

class SmartParseResult(BaseModel):
    """Result of smart parsing"""
    success: bool
    totalRows: int
    parsedRows: int
    customers: List[dict]
    errors: List[str]
    columnMapping: Dict[str, str]
    warnings: List[str]

class BatchSelectionRequest(BaseModel):
    """Request for selective batch creation"""
    customers: List[dict]
    selectedIndices: List[int]
    # Apply to all selected
    subscriptionOverride: Optional[dict] = None

class SubscriptionAction(BaseModel):
    """Subscription management action"""
    action: str  # pause, resume, cancel, update
    subscriptionId: str
    newAmount: Optional[float] = None
    newFrequency: Optional[int] = None

# ==================== COLUMN DETECTION ====================

# Common column name variations
COLUMN_PATTERNS = {
    'firstName': ['first_name', 'firstname', 'first name', 'nombre', 'contact first name', 'name', 'primer nombre'],
    'lastName': ['last_name', 'lastname', 'last name', 'apellido', 'contact last name', 'surname', 'segundo nombre'],
    'email': ['email', 'correo', 'e-mail', 'contact email', 'email address'],
    'phone': ['phone', 'telefono', 'tel', 'mobile', 'celular', 'phone number', 'phone #', 'numero de telefono'],
    'address1': ['address', 'address1', 'direccion', 'street', 'calle', 'address line 1', 'street address'],
    'address2': ['apt', 'apt #', 'apt#', 'suite', 'unit', 'address2', 'address line 2', 'apartment'],
    'city': ['city', 'ciudad', 'town'],
    'state': ['state', 'estado', 'province', 'st'],
    'postalCode': ['zip', 'postal', 'postalcode', 'postal_code', 'codigo_postal', 'zip code', 'zipcode'],
    'country': ['country', 'pais'],
    'checkName': ['check_name', 'checkname', 'account_holder', 'bank account holder name', 'holder name', 'nombre en cuenta'],
    'routing': ['routing', 'routing_number', 'aba', 'bank routing number', 'rtn', 'routing number'],
    'accountNumber': ['account', 'account_number', 'accountnumber', 'bank account number', 'cuenta', 'account number', 'account #'],
    'accountType': ['account_type', 'accounttype', 'tipo_cuenta', 'bank account type', 'type'],
    'accountHolderType': ['holder_type', 'account_holder_type', 'bank account holder type'],
    'planName': ['plan', 'plan_name', 'planname', 'subscription', 'servicio'],
    'amount': ['amount', 'monto', 'price', 'precio', 'plan_amount'],
    'dayFrequency': ['frequency', 'day_frequency', 'frecuencia', 'days', 'dias'],
    'startDate': ['start_date', 'startdate', 'fecha_inicio', 'start'],
    'company': ['company', 'empresa', 'company name', 'business'],
}

def normalize_column_name(col: str) -> str:
    """Normalize column name for matching"""
    return col.lower().strip().replace('_', ' ').replace('-', ' ')

def detect_column_mapping(headers: List[str]) -> Dict[str, str]:
    """
    Detect which CSV column maps to which field
    Returns dict like {'firstName': 'Contact first name', ...}
    """
    mapping = {}
    normalized_headers = {normalize_column_name(h): h for h in headers}
    
    for field, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            normalized_pattern = normalize_column_name(pattern)
            if normalized_pattern in normalized_headers:
                mapping[field] = normalized_headers[normalized_pattern]
                break
    
    return mapping

def extract_zip_from_address(address: str) -> Optional[str]:
    """Try to extract ZIP code from address string"""
    import re
    # Look for 5-digit ZIP
    match = re.search(r'\b(\d{5})(-\d{4})?\b', address)
    if match:
        return match.group(1)
    return None

def clean_phone(phone: str) -> str:
    """Clean phone number to digits only"""
    import re
    return re.sub(r'\D', '', phone)

def clean_routing(routing: str) -> str:
    """Clean routing number"""
    import re
    cleaned = re.sub(r'\D', '', str(routing))
    # Pad with zeros if needed
    return cleaned.zfill(9) if len(cleaned) < 9 else cleaned[:9]

def clean_account(account: str) -> str:
    """Clean account number"""
    import re
    return re.sub(r'\D', '', str(account))

_cached_start_date = None

def get_valid_start_date_str() -> str:
    """Get a valid start date formatted as YYYYMMDD - cached to avoid repeated API calls"""
    global _cached_start_date
    if _cached_start_date:
        return _cached_start_date
    
    try:
        # Try to get real time from worldtimeapi
        import httpx
        response = httpx.get('http://worldtimeapi.org/api/timezone/America/Chicago', timeout=5)
        if response.status_code == 200:
            data = response.json()
            datetime_str = data.get('datetime', '')[:10]
            real_date = datetime.strptime(datetime_str, '%Y-%m-%d')
        else:
            real_date = datetime.now()
    except:
        real_date = datetime.now()
    
    start_date = real_date + timedelta(days=7)
    _cached_start_date = start_date.strftime('%Y%m%d')
    return _cached_start_date

# ==================== SMART CSV PARSER ====================

def smart_parse_csv(
    csv_data: str,
    has_header: bool = True,
    column_hints: Optional[Dict[str, str]] = None,
    default_routing: Optional[str] = None,
    default_amount: Optional[float] = None,
    default_frequency: Optional[int] = None,
    default_plan_name: Optional[str] = None
) -> SmartParseResult:
    """
    Intelligently parse CSV data, auto-detecting columns
    
    Args:
        csv_data: Raw CSV string
        has_header: Whether first row is header
        column_hints: Optional manual column mappings
        default_*: Default values for missing fields
    
    Returns:
        SmartParseResult with parsed customers
    """
    customers = []
    errors = []
    warnings = []
    
    try:
        # Clean CSV data - remove BOM, extra whitespace lines at start
        csv_data = csv_data.strip()
        
        # Skip header lines that are metadata (like "Custom Report - 727 List")
        lines = csv_data.split('\n')
        clean_lines = []
        header_found = False
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            # Detect if this is a data/header line (has tabs or commas with multiple fields)
            tab_count = stripped.count('\t')
            comma_count = stripped.count(',')
            if tab_count >= 3 or comma_count >= 3 or header_found:
                clean_lines.append(line)
                header_found = True
        
        if not clean_lines:
            # Fallback: use all non-empty lines
            clean_lines = [l for l in lines if l.strip()]
        
        csv_data = '\n'.join(clean_lines)
        
        # Auto-detect delimiter: tab vs comma
        first_line = clean_lines[0] if clean_lines else ''
        tab_count = first_line.count('\t')
        comma_count = first_line.count(',')
        
        delimiter = '\t' if tab_count > comma_count else ','
        logger.info(f"Auto-detected delimiter: {'TAB' if delimiter == chr(9) else 'COMMA'} (tabs={tab_count}, commas={comma_count})")
        
        reader = csv.reader(StringIO(csv_data), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return SmartParseResult(
                success=False,
                totalRows=0,
                parsedRows=0,
                customers=[],
                errors=["CSV vacío"],
                columnMapping={},
                warnings=[]
            )
        
        # Detect headers
        if has_header:
            headers = rows[0]
            data_rows = rows[1:]
        else:
            # Generate default headers
            headers = [f"col_{i}" for i in range(len(rows[0]))]
            data_rows = rows
        
        # Auto-detect column mapping
        auto_mapping = detect_column_mapping(headers)
        
        # Apply manual hints if provided
        if column_hints:
            auto_mapping.update(column_hints)
        
        logger.info(f"Column mapping detected: {auto_mapping}")
        
        # Create reverse mapping (CSV column -> field name)
        col_to_field = {v: k for k, v in auto_mapping.items()}
        col_indices = {h: i for i, h in enumerate(headers)}
        
        # Helper to get field value from row
        def get_field(row: List[str], field: str, default: str = '') -> str:
            if field in auto_mapping:
                csv_col = auto_mapping[field]
                if csv_col in col_indices:
                    idx = col_indices[csv_col]
                    if idx < len(row):
                        return row[idx].strip()
            return default
        
        # Process each row
        for row_idx, row in enumerate(data_rows, start=2 if has_header else 1):
            if not any(cell.strip() for cell in row):  # Skip empty rows
                continue
            
            try:
                # Extract customer info
                first_name = get_field(row, 'firstName')
                last_name = get_field(row, 'lastName')
                
                # Try to extract from company name if no first/last name
                if not first_name and not last_name:
                    company_name = get_field(row, 'company')
                    if company_name:
                        parts = company_name.split()
                        if len(parts) >= 2:
                            first_name = parts[0]
                            last_name = ' '.join(parts[1:])
                        else:
                            first_name = company_name
                            last_name = ''
                
                if not first_name:
                    errors.append(f"Fila {row_idx}: Sin nombre")
                    continue
                
                # Build customer object
                email = get_field(row, 'email')
                phone = clean_phone(get_field(row, 'phone'))
                address = get_field(row, 'address1')
                city = get_field(row, 'city')
                state = get_field(row, 'state')
                postal_code = get_field(row, 'postalCode')
                country = get_field(row, 'country', 'US')
                
                # Try to extract ZIP from address if missing
                if not postal_code and address:
                    extracted_zip = extract_zip_from_address(address)
                    if extracted_zip:
                        postal_code = extracted_zip
                        warnings.append(f"Fila {row_idx}: ZIP extraído de dirección")
                
                # Fix country if it looks like phone
                if country and len(country) > 3:
                    country = 'US'
                
                # Bank info
                check_name = get_field(row, 'checkName')
                if not check_name:
                    check_name = f"{first_name} {last_name}".strip()
                
                routing = clean_routing(get_field(row, 'routing'))
                if not routing or routing == '000000000':
                    if default_routing:
                        routing = default_routing
                        warnings.append(f"Fila {row_idx}: Usando routing predeterminado")
                    else:
                        routing = ''
                        warnings.append(f"Fila {row_idx}: Sin routing number - requerido para Merchant One")
                
                if routing and len(routing) != 9:
                    warnings.append(f"Fila {row_idx}: Routing debe ser 9 dígitos (tiene {len(routing)})")
                    routing = ''
                
                account_number = clean_account(get_field(row, 'accountNumber'))
                if not account_number:
                    warnings.append(f"Fila {row_idx}: Sin número de cuenta - requerido para Merchant One")
                
                account_type = get_field(row, 'accountType', 'checking').lower()
                if account_type not in ['checking', 'savings']:
                    account_type = 'checking'
                
                account_holder_type = get_field(row, 'accountHolderType', 'personal').lower()
                if 'business' in account_holder_type or 'negocio' in account_holder_type:
                    account_holder_type = 'business'
                else:
                    account_holder_type = 'personal'
                
                # Subscription info
                plan_name = get_field(row, 'planName')
                if not plan_name:
                    plan_name = default_plan_name or 'Plan Mensual'
                
                amount_str = get_field(row, 'amount')
                try:
                    amount = float(amount_str.replace('$', '').replace(',', '')) if amount_str else (default_amount or 50.0)
                except:
                    amount = default_amount or 50.0
                
                frequency_str = get_field(row, 'dayFrequency')
                try:
                    frequency = int(frequency_str) if frequency_str else (default_frequency or 30)
                except:
                    frequency = default_frequency or 30
                
                start_date = get_field(row, 'startDate')
                if not start_date:
                    start_date = get_valid_start_date_str()
                else:
                    # Try to normalize date format to YYYYMMDD
                    start_date = start_date.replace('-', '').replace('/', '')
                    if len(start_date) == 8:
                        # Check if MMDDYYYY and convert to YYYYMMDD
                        if int(start_date[:2]) <= 12:  # Looks like MMDDYYYY
                            start_date = start_date[4:8] + start_date[0:4]
                    else:
                        start_date = get_valid_start_date_str()
                
                # Build customer object
                customer = {
                    "customer": {
                        "firstName": first_name,
                        "lastName": last_name or '',
                        "email": email,
                        "phone": phone,
                        "company": get_field(row, 'company', ''),
                        "address1": address,
                        "address2": "",
                        "city": city,
                        "state": state[:2].upper() if state else '',
                        "postalCode": postal_code,
                        "country": "US"
                    },
                    "bank": {
                        "checkName": check_name,
                        "routing": routing,
                        "accountNumber": account_number,
                        "accountHolderType": account_holder_type,
                        "accountType": account_type,
                        "secCode": "PPD" if account_holder_type == 'personal' else "CCD"
                    },
                    "subscription": {
                        "planName": plan_name,
                        "amount": amount,
                        "dayFrequency": frequency,
                        "startDate": start_date,
                        "planPayments": 0,
                        "productSku": "",
                        "orderDescription": ""
                    },
                    "useAchDefaults": True,
                    "_sourceRow": row_idx
                }
                
                customers.append(customer)
                
            except Exception as e:
                errors.append(f"Fila {row_idx}: Error procesando - {str(e)}")
        
        return SmartParseResult(
            success=len(customers) > 0,
            totalRows=len(data_rows),
            parsedRows=len(customers),
            customers=customers,
            errors=errors,
            columnMapping=auto_mapping,
            warnings=warnings
        )
        
    except Exception as e:
        logger.error(f"Smart parse error: {e}")
        return SmartParseResult(
            success=False,
            totalRows=0,
            parsedRows=0,
            customers=[],
            errors=[f"Error parseando CSV: {str(e)}"],
            columnMapping={},
            warnings=[]
        )


# ==================== SUBSCRIPTION MANAGEMENT ====================

MERCHANT_ONE_API_URL = os.getenv(
    'MERCHANTONE_API_URL',
    'https://secure.networkmerchants.com/api/transact.php'
)
MERCHANT_ONE_SECURITY_KEY = os.getenv('MERCHANTONE_PRIVATE_SECURITY_KEY', '')

async def pause_subscription(subscription_id: str) -> Dict[str, Any]:
    """Pause a recurring subscription"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'recurring': 'pause_subscription',
        'subscription_id': subscription_id
    }
    return await _make_nmi_request(payload)

async def resume_subscription(subscription_id: str) -> Dict[str, Any]:
    """Resume a paused subscription"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'recurring': 'resume_subscription',
        'subscription_id': subscription_id
    }
    return await _make_nmi_request(payload)

async def cancel_subscription(subscription_id: str) -> Dict[str, Any]:
    """Cancel/delete a subscription"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'recurring': 'delete_subscription',
        'subscription_id': subscription_id
    }
    return await _make_nmi_request(payload)

async def update_subscription(
    subscription_id: str,
    new_amount: Optional[float] = None,
    new_frequency: Optional[int] = None
) -> Dict[str, Any]:
    """Update subscription amount or frequency"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'recurring': 'update_subscription',
        'subscription_id': subscription_id
    }
    
    if new_amount is not None:
        payload['plan_amount'] = f"{new_amount:.2f}"
    if new_frequency is not None:
        payload['day_frequency'] = str(new_frequency)
    
    return await _make_nmi_request(payload)

async def delete_vault_customer(customer_vault_id: str) -> Dict[str, Any]:
    """Delete customer from vault"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'customer_vault': 'delete_customer',
        'customer_vault_id': customer_vault_id
    }
    return await _make_nmi_request(payload)


async def update_vault_customer(
    customer_vault_id: str,
    # Customer info
    first_name: Optional[str] = None,
    last_name: Optional[str] = None,
    address1: Optional[str] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    zip_code: Optional[str] = None,
    email: Optional[str] = None,
    phone: Optional[str] = None,
    # Bank info
    check_name: Optional[str] = None,
    check_aba: Optional[str] = None,  # Routing number
    check_account: Optional[str] = None,  # Account number
    account_type: Optional[str] = None,  # checking/savings
    account_holder_type: Optional[str] = None  # personal/business
) -> Dict[str, Any]:
    """
    Update customer vault information including bank details
    
    NMI API: customer_vault=update_customer
    """
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'customer_vault': 'update_customer',
        'customer_vault_id': customer_vault_id
    }
    
    # Add customer info if provided
    if first_name:
        payload['first_name'] = first_name
    if last_name:
        payload['last_name'] = last_name
    if address1:
        payload['address1'] = address1
    if city:
        payload['city'] = city
    if state:
        payload['state'] = state
    if zip_code:
        payload['zip'] = zip_code
    if email:
        payload['email'] = email
    if phone:
        payload['phone'] = phone
    
    # Add bank info if provided
    if check_name:
        payload['checkname'] = check_name
    if check_aba:
        payload['checkaba'] = check_aba
    if check_account:
        payload['checkaccount'] = check_account
    if account_type:
        payload['account_type'] = account_type
    if account_holder_type:
        payload['account_holder_type'] = account_holder_type
    
    return await _make_nmi_request(payload)


async def charge_vault_customer(
    customer_vault_id: str,
    amount: float,
    order_description: Optional[str] = None
) -> Dict[str, Any]:
    """Process a one-time charge for a vault customer"""
    from merchant_one_service import get_descriptor_fields
    
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'type': 'sale',
        'customer_vault_id': customer_vault_id,
        'amount': f"{amount:.2f}"
    }
    
    if order_description:
        payload['order_description'] = order_description
    
    # Add statement descriptor
    payload.update(get_descriptor_fields())
    
    return await _make_nmi_request(payload)

async def _make_nmi_request(payload: Dict[str, str]) -> Dict[str, Any]:
    """Make request to NMI API"""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                MERCHANT_ONE_API_URL,
                data=payload,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            
            # Parse response
            from urllib.parse import parse_qs
            parsed = parse_qs(response.text, keep_blank_values=True)
            data = {k: v[0] if len(v) == 1 else v for k, v in parsed.items()}
            
            success = data.get('response') == '1'
            
            return {
                'success': success,
                'responseCode': data.get('response'),
                'responseText': data.get('responsetext', ''),
                'transactionId': data.get('transactionid'),
                'rawResponse': response.text
            }
            
    except Exception as e:
        logger.error(f"NMI request error: {e}")
        return {
            'success': False,
            'error': str(e)
        }


# ==================== CSV TEMPLATE GENERATOR ====================

def generate_csv_template(include_example: bool = True) -> str:
    """Generate CSV template with headers and optional example row"""
    headers = [
        'firstName',
        'lastName', 
        'email',
        'phone',
        'address1',
        'city',
        'state',
        'postalCode',
        'checkName',
        'routing',
        'accountNumber',
        'accountType',
        'planName',
        'amount',
        'dayFrequency',
        'startDate'
    ]
    
    csv_content = ','.join(headers) + '\n'
    
    if include_example:
        example = [
            'Juan',
            'Pérez García',
            'juan.perez@email.com',
            '(832) 555-1234',
            '123 Main Street Apt 4B',
            'Houston',
            'TX',
            '77001',
            'Juan Pérez García',
            '111000025',
            '123456789',
            'checking',
            'Plan Mensual',
            '50.00',
            '30',
            '20250401'
        ]
        csv_content += ','.join(example) + '\n'
        
        # Add second example
        example2 = [
            'María',
            'González López',
            'maria.gonzalez@email.com',
            '(713) 555-5678',
            '456 Oak Avenue',
            'Dallas',
            'TX',
            '75201',
            'María González López',
            '111000025',
            '987654321',
            'savings',
            'Plan Semanal',
            '25.00',
            '7',
            '20250408'
        ]
        csv_content += ','.join(example2)
    
    return csv_content


def generate_excel_template_base64() -> str:
    """Generate Excel template and return as base64"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        import base64
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes ACH"
        
        # Headers
        headers = [
            ('A', 'firstName', 'Nombre'),
            ('B', 'lastName', 'Apellido'),
            ('C', 'email', 'Email'),
            ('D', 'phone', 'Teléfono'),
            ('E', 'address1', 'Dirección'),
            ('F', 'city', 'Ciudad'),
            ('G', 'state', 'Estado'),
            ('H', 'postalCode', 'Código Postal'),
            ('I', 'checkName', 'Nombre en Cuenta'),
            ('J', 'routing', 'Routing (9 dígitos)'),
            ('K', 'accountNumber', 'Número de Cuenta'),
            ('L', 'accountType', 'Tipo (checking/savings)'),
            ('M', 'planName', 'Nombre del Plan'),
            ('N', 'amount', 'Monto $'),
            ('O', 'dayFrequency', 'Frecuencia (días)'),
            ('P', 'startDate', 'Fecha Inicio (YYYYMMDD)')
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, field, display in headers:
            cell = ws[f"{col}1"]
            cell.value = display
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[col].width = 18
        
        # Example rows
        examples = [
            ['Juan', 'Pérez García', 'juan@email.com', '(832)555-1234', '123 Main St', 'Houston', 'TX', '77001', 'Juan Pérez García', '111000025', '123456789', 'checking', 'Plan Mensual', '50.00', '30', '20250401'],
            ['María', 'González', 'maria@email.com', '(713)555-5678', '456 Oak Ave', 'Dallas', 'TX', '75201', 'María González', '111000025', '987654321', 'checking', 'Plan Mensual', '50.00', '30', '20250401'],
        ]
        
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, value in enumerate(example):
                cell = ws.cell(row=row_idx, column=col_idx+1, value=value)
                cell.border = thin_border
        
        # Add instructions sheet
        ws2 = wb.create_sheet(title="Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA LLENAR EL TEMPLATE",
            "",
            "CAMPOS OBLIGATORIOS (*)",
            "• firstName: Nombre del cliente",
            "• lastName: Apellido del cliente", 
            "• address1: Dirección completa",
            "• city: Ciudad",
            "• state: Estado (2 letras, ej: TX, OK, CA)",
            "• postalCode: Código postal (5 dígitos)",
            "• routing: Número de routing del banco (9 dígitos)",
            "• accountNumber: Número de cuenta bancaria",
            "",
            "CAMPOS OPCIONALES",
            "• email: Correo electrónico",
            "• phone: Teléfono (cualquier formato)",
            "• checkName: Nombre como aparece en cheques (si no se pone, usa nombre+apellido)",
            "• accountType: 'checking' o 'savings' (default: checking)",
            "• planName: Nombre del plan de suscripción",
            "• amount: Monto a cobrar (ej: 50.00)",
            "• dayFrequency: Cada cuántos días cobrar (7=semanal, 30=mensual)",
            "• startDate: Fecha de inicio en formato YYYYMMDD (ej: 20250401)",
            "",
            "NOTAS IMPORTANTES",
            "• El routing number DEBE ser 9 dígitos",
            "• Si todos los clientes usan el mismo banco, puedes dejar routing vacío y configurarlo en la app",
            "• Puedes importar desde otro Excel copiando y pegando las columnas",
            "• La app detectará automáticamente las columnas si tienen nombres similares",
        ]
        
        for row_idx, text in enumerate(instructions, start=1):
            ws2.cell(row=row_idx, column=1, value=text)
            if row_idx == 1:
                ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
            elif text.startswith("CAMPOS") or text.startswith("NOTAS"):
                ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        
        ws2.column_dimensions['A'].width = 80
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return base64.b64encode(buffer.read()).decode('utf-8')
        
    except ImportError:
        logger.warning("openpyxl not installed, returning CSV format")
        return None



# ==================== QUERY API - TRANSACTION HISTORY ====================

MERCHANT_ONE_QUERY_URL = os.getenv(
    'MERCHANTONE_QUERY_URL',
    'https://secure.networkmerchants.com/api/query.php'
)


async def query_transactions(
    customer_vault_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    condition: Optional[str] = None,
    limit: int = 50
) -> Dict[str, Any]:
    """
    Query transaction history from NMI
    
    Args:
        customer_vault_id: Filter by vault customer
        start_date: Start date (YYYYMMDD)
        end_date: End date (YYYYMMDD)
        transaction_type: sale, auth, credit, validate, offline
        condition: pending, pendingsettlement, failed, canceled, complete, unknown
        limit: Max results
    
    Returns:
        Dict with transactions list
    """
    import xml.etree.ElementTree as ET
    
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
    }
    
    if customer_vault_id:
        payload['customer_vault_id'] = customer_vault_id
    if start_date:
        payload['start_date'] = start_date
    if end_date:
        payload['end_date'] = end_date
    if transaction_type:
        payload['transaction_type'] = transaction_type
    if condition:
        payload['condition'] = condition
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                MERCHANT_ONE_QUERY_URL,
                data=payload,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            
            # Parse XML response
            transactions = []
            try:
                root = ET.fromstring(response.text)
                
                for trans in root.findall('.//transaction'):
                    transaction = {
                        'transactionId': trans.findtext('transaction_id', ''),
                        'amount': trans.findtext('amount', '0'),
                        'condition': trans.findtext('condition', ''),
                        'transactionType': trans.findtext('transaction_type', ''),
                        'date': trans.findtext('date', ''),
                        'time': trans.findtext('time', ''),
                        'firstName': trans.findtext('first_name', ''),
                        'lastName': trans.findtext('last_name', ''),
                        'customerVaultId': trans.findtext('customer_vault_id', ''),
                        'responseCode': trans.findtext('response_code', ''),
                        'responseText': trans.findtext('response_text', ''),
                    }
                    transactions.append(transaction)
                    
                    if len(transactions) >= limit:
                        break
                
            except ET.ParseError:
                logger.error(f"Failed to parse XML response: {response.text[:500]}")
            
            return {
                'success': True,
                'transactions': transactions,
                'count': len(transactions)
            }
            
    except Exception as e:
        logger.error(f"Query API error: {e}")
        return {
            'success': False,
            'error': str(e),
            'transactions': []
        }


async def get_subscription_info(subscription_id: str) -> Dict[str, Any]:
    """Get subscription details from NMI"""
    payload = {
        'security_key': MERCHANT_ONE_SECURITY_KEY,
        'report_type': 'recurring',
        'subscription_id': subscription_id
    }
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                MERCHANT_ONE_QUERY_URL,
                data=payload,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            
            import xml.etree.ElementTree as ET
            
            try:
                root = ET.fromstring(response.text)
                
                sub = root.find('.//subscription')
                if sub is not None:
                    return {
                        'success': True,
                        'subscription': {
                            'subscriptionId': sub.findtext('subscription_id', ''),
                            'planAmount': sub.findtext('plan_amount', '0'),
                            'planName': sub.findtext('plan_name', ''),
                            'dayFrequency': sub.findtext('day_frequency', ''),
                            'startDate': sub.findtext('start_date', ''),
                            'nextChargeDate': sub.findtext('next_charge_date', ''),
                            'status': sub.findtext('status', ''),
                            'completedPayments': sub.findtext('completed_payments', '0'),
                            'remainingPayments': sub.findtext('remaining_payments', ''),
                        }
                    }
                    
            except ET.ParseError:
                pass
            
            return {
                'success': False,
                'error': 'Subscription not found'
            }
            
    except Exception as e:
        logger.error(f"Get subscription error: {e}")
        return {
            'success': False,
            'error': str(e)
        }


# ==================== EXPORT FUNCTIONS ====================

def generate_customers_csv(customers: List[dict]) -> str:
    """Generate CSV export of customers"""
    from io import StringIO
    import csv
    
    output = StringIO()
    
    fieldnames = [
        'Nombre', 'Apellido', 'Email', 'Teléfono',
        'Cuenta (enmascarada)', 'Vault ID', 'Plan', 'Monto',
        'Frecuencia (días)', 'Estado Suscripción', 'Fecha Creación'
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for c in customers:
        writer.writerow({
            'Nombre': c.get('firstName', ''),
            'Apellido': c.get('lastName', ''),
            'Email': c.get('email', ''),
            'Teléfono': c.get('phone', ''),
            'Cuenta (enmascarada)': c.get('maskedAccount', ''),
            'Vault ID': c.get('customerVaultId', ''),
            'Plan': c.get('planName', ''),
            'Monto': c.get('planAmount', ''),
            'Frecuencia (días)': c.get('dayFrequency', ''),
            'Estado Suscripción': c.get('subscriptionStatus', 'none'),
            'Fecha Creación': c.get('createdAt', ''),
        })
    
    return output.getvalue()


def generate_customers_excel(customers: List[dict]) -> Optional[str]:
    """Generate Excel export of customers"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        import base64
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes Vault"
        
        # Headers
        headers = [
            'Nombre', 'Apellido', 'Email', 'Teléfono',
            'Cuenta', 'Vault ID', 'Plan', 'Monto ($)',
            'Frecuencia', 'Estado', 'Fecha Creación'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row_idx, c in enumerate(customers, 2):
            data = [
                c.get('firstName', ''),
                c.get('lastName', ''),
                c.get('email', ''),
                c.get('phone', ''),
                c.get('maskedAccount', ''),
                c.get('customerVaultId', ''),
                c.get('planName', ''),
                c.get('planAmount', ''),
                c.get('dayFrequency', ''),
                c.get('subscriptionStatus', 'none'),
                str(c.get('createdAt', ''))[:19] if c.get('createdAt') else '',
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color by status
                if col_idx == 10:  # Status column
                    if value == 'active':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == 'paused':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == 'cancelled':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return base64.b64encode(buffer.read()).decode('utf-8')
        
    except ImportError:
        return None


# ==================== BULK ACTIONS ====================

async def bulk_pause_subscriptions(subscription_ids: List[str]) -> Dict[str, Any]:
    """Pause multiple subscriptions"""
    results = []
    success_count = 0
    fail_count = 0
    
    for sub_id in subscription_ids:
        result = await pause_subscription(sub_id)
        results.append({
            'subscriptionId': sub_id,
            'success': result.get('success', False),
            'error': result.get('responseText') if not result.get('success') else None
        })
        
        if result.get('success'):
            success_count += 1
        else:
            fail_count += 1
    
    return {
        'success': fail_count == 0,
        'totalProcessed': len(subscription_ids),
        'successCount': success_count,
        'failCount': fail_count,
        'results': results
    }


async def bulk_resume_subscriptions(subscription_ids: List[str]) -> Dict[str, Any]:
    """Resume multiple subscriptions"""
    results = []
    success_count = 0
    fail_count = 0
    
    for sub_id in subscription_ids:
        result = await resume_subscription(sub_id)
        results.append({
            'subscriptionId': sub_id,
            'success': result.get('success', False),
            'error': result.get('responseText') if not result.get('success') else None
        })
        
        if result.get('success'):
            success_count += 1
        else:
            fail_count += 1
    
    return {
        'success': fail_count == 0,
        'totalProcessed': len(subscription_ids),
        'successCount': success_count,
        'failCount': fail_count,
        'results': results
    }


async def bulk_cancel_subscriptions(subscription_ids: List[str]) -> Dict[str, Any]:
    """Cancel multiple subscriptions"""
    results = []
    success_count = 0
    fail_count = 0
    
    for sub_id in subscription_ids:
        result = await cancel_subscription(sub_id)
        results.append({
            'subscriptionId': sub_id,
            'success': result.get('success', False),
            'error': result.get('responseText') if not result.get('success') else None
        })
        
        if result.get('success'):
            success_count += 1
        else:
            fail_count += 1
    
    return {
        'success': fail_count == 0,
        'totalProcessed': len(subscription_ids),
        'successCount': success_count,
        'failCount': fail_count,
        'results': results
    }
