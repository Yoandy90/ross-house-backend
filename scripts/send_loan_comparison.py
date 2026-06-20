"""Loan Comparison Analysis + Excel - 5 lenders side by side."""
import os, base64
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
load_dotenv(Path('/app/ross-house-backend/.env'))

from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, HRFlowable)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName,
                                    FileType, Disposition)

NAVY = colors.HexColor("#0E3A66"); ACCENT = colors.HexColor("#0E5AA7")
GREEN = colors.HexColor("#16A34A"); GREEN_BG = colors.HexColor("#ECFDF5")
RED = colors.HexColor("#DC2626"); RED_BG = colors.HexColor("#FEF2F2")
AMBER = colors.HexColor("#D97706"); AMBER_BG = colors.HexColor("#FEF3C7")
GRAY = colors.HexColor("#6B7280"); DARK = colors.HexColor("#1F2937")
LIGHT_BG = colors.HexColor("#F0F4F8"); BORDER = colors.HexColor("#D1D5DB")

def pmt(loan, rate, years):
    m = rate / 12; n = years * 12
    return loan * (m * (1 + m) ** n) / ((1 + m) ** n - 1) if loan > 0 else 0

# Lenders setup
LENDERS = [
    # (name, loan_amount, rate, amort_years, term_years, recourse, notes)
    ("Freddie Mac SBL", 5625000, 0.0625, 30, 10, "Non-recourse", "Best for stabilized assets"),
    ("Fannie Mae DUS Small", 5625000, 0.0650, 30, 10, "Non-recourse", "Similar to Freddie"),
    ("CMBS (Walker&Dunlop)", 5625000, 0.0675, 30, 10, "Non-recourse", "Conduit loan"),
    ("HUD 223(f)", 6375000, 0.0600, 35, 35, "Non-recourse", "Highest LTV 85%, 6-12mo close"),
    ("Amarillo Nat'l Bank", 5250000, 0.0725, 25, 5, "RECOURSE", "Local TX, balloon at yr 5"),
]
SELLER_NOTE_PMT = 750000 * 0.06 / 12  # interest-only

def build_pdf(path):
    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=0.55*inch,
                             rightMargin=0.55*inch, topMargin=0.6*inch,
                             bottomMargin=0.6*inch,
                             title="Jasmine Loan Comparison Analysis")
    base = getSampleStyleSheet()
    title = ParagraphStyle("T", parent=base["Title"], fontSize=18, textColor=NAVY,
                            alignment=TA_CENTER, spaceAfter=4)
    sub = ParagraphStyle("S", parent=base["Normal"], fontSize=10, textColor=GRAY,
                          alignment=TA_CENTER, spaceAfter=12)
    h2 = ParagraphStyle("H2", parent=base["Heading2"], fontSize=13, textColor=NAVY,
                         spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold")
    h3 = ParagraphStyle("H3", parent=base["Heading3"], fontSize=11, textColor=ACCENT,
                         spaceBefore=8, spaceAfter=4, fontName="Helvetica-Bold")
    body = ParagraphStyle("B", parent=base["Normal"], fontSize=9.5, leading=13,
                           textColor=DARK)
    small = ParagraphStyle("Sm", parent=base["Normal"], fontSize=8.5, textColor=GRAY,
                            leading=11, alignment=TA_CENTER)

    story = []
    story.append(Paragraph("JASMINE APARTMENTS - LOAN COMPARISON ANALYSIS", title))
    story.append(Paragraph("5 Senior Debt Options + Seller Note - Side by Side", sub))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=12))

    # Income scenario
    story.append(Paragraph("Income Assumptions", h2))
    income = [
        ["Scenario", "Avg Rent/Unit", "Gross Income/yr", "NOI/yr (50% opex)", "NOI/mo"],
        ["Conservative", "$600", "$1,022,400", "$475,416", "$39,618"],
        ["Realistic (Joe's listings)", "$675", "$1,150,200", "$534,843", "$44,570"],
        ["Optimized (45% opex)", "$675", "$1,150,200", "$588,327", "$49,027"],
    ]
    it = Table(income, colWidths=[2.0*inch, 1.2*inch, 1.4*inch, 1.5*inch, 1.1*inch])
    it.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                             ("BACKGROUND",(0,3),(-1,3),GREEN_BG),
                             ("GRID",(0,0),(-1,-1),0.4,BORDER),
                             ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                             ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                             ("FONTSIZE",(0,0),(-1,-1),9),
                             ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(it)
    story.append(Spacer(1, 10))

    # Senior debt comparison
    story.append(Paragraph("5 Senior Debt Options Comparison", h2))
    rows = [["Lender", "Loan", "Rate", "Amort", "Mo Payment", "Annual Pmt", "Recourse"]]
    for name, loan, rate, amort, term, recourse, notes in LENDERS:
        p = pmt(loan, rate, amort)
        rows.append([name, f"${loan/1e6:.3f}M", f"{rate*100:.2f}%", f"{amort}yr",
                      f"${p:,.0f}", f"${p*12:,.0f}", recourse])
    t = Table(rows, colWidths=[1.8*inch, 0.9*inch, 0.7*inch, 0.65*inch, 1.0*inch, 1.0*inch, 1.05*inch])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                            ("BACKGROUND",(0,1),(-1,1),GREEN_BG),  # Freddie row highlight
                            ("GRID",(0,0),(-1,-1),0.4,BORDER),
                            ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                            ("FONTSIZE",(0,0),(-1,-1),8.5),
                            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(t)
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "<i>Note: Seller note $750K @ 6% IO adds $3,750/mo (interest only) for all scenarios.</i>",
        small))
    story.append(Spacer(1, 12))

    # Cash flow per scenario
    for scenario_name, noi_monthly in [
        ("Scenario 1: Conservative ($600 rent, 50% opex)", 39618),
        ("Scenario 2: Realistic ($675 rent, 50% opex)", 44570),
        ("Scenario 3: Optimized ($675 rent, 45% opex)", 49027),
    ]:
        story.append(Paragraph(scenario_name, h3))
        rows_cf = [["Lender", "Senior Pmt", "Seller IO", "Total Pmt", "NOI", "Cash Flow", "DSCR"]]
        for name, loan, rate, amort, term, recourse, notes in LENDERS:
            sp = pmt(loan, rate, amort)
            total = sp + SELLER_NOTE_PMT
            cf = noi_monthly - total
            dscr = noi_monthly / total if total else 0
            cf_text = f"${cf:+,.0f}"
            rows_cf.append([name, f"${sp:,.0f}", f"${SELLER_NOTE_PMT:,.0f}",
                             f"${total:,.0f}", f"${noi_monthly:,.0f}",
                             cf_text, f"{dscr:.2f}x"])
        t2 = Table(rows_cf, colWidths=[1.8*inch, 0.95*inch, 0.8*inch, 0.95*inch,
                                          0.85*inch, 1.0*inch, 0.7*inch])
        styles_t = [("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                    ("GRID",(0,0),(-1,-1),0.4,BORDER),
                    ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                    ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                    ("FONTSIZE",(0,0),(-1,-1),8.2),
                    ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]
        # color rows based on DSCR
        for i, (name, loan, rate, amort, term, _, _) in enumerate(LENDERS, 1):
            sp = pmt(loan, rate, amort)
            total = sp + SELLER_NOTE_PMT
            dscr = noi_monthly / total if total else 0
            if dscr >= 1.20:
                styles_t.append(("BACKGROUND", (5, i), (6, i), GREEN_BG))
            elif dscr >= 1.10:
                styles_t.append(("BACKGROUND", (5, i), (6, i), AMBER_BG))
            else:
                styles_t.append(("BACKGROUND", (5, i), (6, i), RED_BG))
        t2.setStyle(TableStyle(styles_t))
        story.append(t2)
        story.append(Spacer(1, 8))

    story.append(PageBreak())

    # Recommendation
    story.append(Paragraph("Recommendation: Freddie Mac SBL", h2))

    win = [
        ["Reason", "Detail"],
        ["Lowest payment", "$34,627/mo (vs $37,939 bank = $3,312/mo savings = $1.2M over 30 yrs)"],
        ["Best rate", "6.25% fixed for 10 years"],
        ["Non-recourse", "You don't sign personal guarantee - LLC is solely liable"],
        ["30-year amortization", "Lower payment vs bank's 25 years"],
        ["Quick close", "45-60 days (vs HUD 6-12 months)"],
        ["TX-friendly", "Major SBL volume in Texas tertiary markets"],
        ["LTV", "75% standard (lets you keep cash reserves)"],
    ]
    wt = Table(win, colWidths=[2.0*inch, 5.0*inch])
    wt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),
                             ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                             ("BACKGROUND",(0,1),(0,-1),LIGHT_BG),
                             ("GRID",(0,0),(-1,-1),0.4,BORDER),
                             ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                             ("FONTSIZE",(0,0),(-1,-1),9),
                             ("LEFTPADDING",(0,0),(-1,-1),8),
                             ("RIGHTPADDING",(0,0),(-1,-1),8),
                             ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    story.append(wt)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Final Recommended Cash Flow (Freddie SBL + Seller Note)", h2))
    final = [
        ["", "Conservative", "Realistic", "Optimized"],
        ["Monthly NOI", "$39,618", "$44,570", "$49,027"],
        ["Freddie SBL payment", "$34,627", "$34,627", "$34,627"],
        ["Seller note IO", "$3,750", "$3,750", "$3,750"],
        ["Total monthly debt", "$38,377", "$38,377", "$38,377"],
        ["MONTHLY CASH FLOW", "+$1,241", "+$6,193", "+$10,650"],
        ["ANNUAL CASH FLOW", "+$14,892", "+$74,316", "+$127,800"],
        ["DSCR", "1.03x", "1.16x", "1.28x"],
    ]
    ft = Table(final, colWidths=[2.2*inch, 1.6*inch, 1.6*inch, 1.6*inch])
    ft.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),
                             ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                             ("BACKGROUND",(0,5),(-1,7),GREEN_BG),
                             ("FONTNAME",(0,5),(-1,7),"Helvetica-Bold"),
                             ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
                             ("GRID",(0,0),(-1,-1),0.4,BORDER),
                             ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                             ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                             ("FONTSIZE",(0,0),(-1,-1),10),
                             ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    story.append(ft)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Critical Action Items", h2))
    for it_text in [
        "Confirm with Joe Kuruvila the ACTUAL average rent (likely $650-700, not $600)",
        "Negotiate seller financing aggressively - the $3,750/mo IO is key to making this cashflow",
        "Lock Freddie SBL rate as soon as possible - rates volatile in 2026",
        "If Joe refuses seller financing, increase your offer to $7.2M instead of $7.5M",
        "Underwrite OpEx carefully - $475K is conservative, actual could be $440-470K",
        "Build $200K capex reserve at closing (not in the $475K NOI)",
    ]:
        story.append(Paragraph(f"- {it_text}", body))
    story.append(Spacer(1, 14))

    story.append(Paragraph(
        f"Ross House Rentals LLC - Loan Comparison Analysis - "
        f"{datetime.now().strftime('%B %d, %Y')}", small))
    doc.build(story)


def build_excel(path):
    wb = Workbook()
    NAVY_F = PatternFill("solid", fgColor="0E3A66")
    GREEN_F = PatternFill("solid", fgColor="ECFDF5")
    AMBER_F = PatternFill("solid", fgColor="FEF3C7")
    RED_F = PatternFill("solid", fgColor="FEF2F2")
    LIGHT_F = PatternFill("solid", fgColor="F0F4F8")
    WHITE = Font(color="FFFFFF", bold=True, size=11)
    BOLD = Font(bold=True, size=10)
    RIGHT = Alignment(horizontal="right", vertical="center")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.active
    ws.title = "Loan_Comparison"
    ws.column_dimensions["A"].width = 28
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18

    ws["A1"] = "JASMINE - LOAN COMPARISON (5 LENDERS) - 142 UNITS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = NAVY_F
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Inputs section
    inputs = [
        ("", "", "", "", "", ""),
        ("INPUT ASSUMPTIONS (editable)", "", "", "", "", ""),
        ("Total units", 142, "", "", "", ""),
        ("Average rent per unit/month", 675, "← Edit this", "", "", ""),
        ("Annual vacancy", 0.07, "", "", "", ""),
        ("Operating expense ratio", 0.50, "← Edit this (45-55%)", "", "", ""),
        ("Seller note balance", 750000, "", "", "", ""),
        ("Seller note rate (IO)", 0.06, "", "", "", ""),
    ]
    for i, row in enumerate(inputs, 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val if val != "" else None)
            if j == 1 and isinstance(val, str) and val and row[1] == "":
                c.font = Font(color="0E3A66", bold=True, size=11)
                c.fill = LIGHT_F
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            elif j == 1 and isinstance(val, str) and val:
                c.font = BOLD
        if isinstance(row[1], float) and row[1] < 1:
            ws.cell(row=i, column=2).number_format = "0.00%"
        elif isinstance(row[1], (int, float)) and row[1] != "" and row[1] is not None:
            ws.cell(row=i, column=2).number_format = "#,##0"

    # Formulas for NOI
    # Row 10 is empty
    ws.cell(row=11, column=1, value="CALCULATED NOI").font = Font(color="0E3A66", bold=True)
    ws.cell(row=11, column=1).fill = LIGHT_F
    ws.merge_cells("A11:F11")

    ws.cell(row=12, column=1, value="Gross potential rent (annual)").font = BOLD
    ws.cell(row=12, column=2, value="=B3*B4*12")
    ws.cell(row=12, column=2).number_format = "$#,##0"

    ws.cell(row=13, column=1, value="Less: Vacancy").font = BOLD
    ws.cell(row=13, column=2, value="=-B12*B5")
    ws.cell(row=13, column=2).number_format = "$#,##0"

    ws.cell(row=14, column=1, value="Effective Gross Income").font = BOLD
    ws.cell(row=14, column=2, value="=B12+B13")
    ws.cell(row=14, column=2).number_format = "$#,##0"

    ws.cell(row=15, column=1, value="Less: Operating expenses").font = BOLD
    ws.cell(row=15, column=2, value="=-B14*B6")
    ws.cell(row=15, column=2).number_format = "$#,##0"

    ws.cell(row=16, column=1, value="ANNUAL NOI").font = Font(bold=True, color="065F46", size=11)
    ws.cell(row=16, column=1).fill = GREEN_F
    ws.cell(row=16, column=2, value="=B14+B15")
    ws.cell(row=16, column=2).number_format = "$#,##0"
    ws.cell(row=16, column=2).fill = GREEN_F
    ws.cell(row=16, column=2).font = Font(bold=True, color="065F46", size=11)

    ws.cell(row=17, column=1, value="MONTHLY NOI").font = BOLD
    ws.cell(row=17, column=2, value="=B16/12")
    ws.cell(row=17, column=2).number_format = "$#,##0"

    # Loan comparison table
    ws.cell(row=19, column=1, value="LENDER COMPARISON").font = Font(color="0E3A66", bold=True, size=12)
    ws.cell(row=19, column=1).fill = LIGHT_F
    ws.merge_cells("A19:F19")

    headers = ["Metric", "Freddie SBL", "Fannie DUS", "CMBS", "HUD 223(f)", "Bank"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=20, column=j, value=h)
        c.font = WHITE
        c.fill = NAVY_F
        c.alignment = CENTER

    # Loan data per column
    loan_amounts = [5625000, 5625000, 5625000, 6375000, 5250000]
    rates = [0.0625, 0.0650, 0.0675, 0.0600, 0.0725]
    amorts = [30, 30, 30, 35, 25]

    metrics = [
        ("Loan amount", loan_amounts, "$#,##0"),
        ("Interest rate", rates, "0.00%"),
        ("Amortization (years)", amorts, "0"),
    ]
    for i, (label, vals, fmt) in enumerate(metrics, 21):
        ws.cell(row=i, column=1, value=label).font = BOLD
        for j, v in enumerate(vals, 2):
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = fmt
            c.alignment = RIGHT

    # Monthly payment formula row 24
    ws.cell(row=24, column=1, value="Monthly payment (P&I)").font = BOLD
    for j in range(2, 7):
        col = get_column_letter(j)
        # PMT formula
        formula = f'=PMT({col}22/12,{col}23*12,-{col}21)'
        c = ws.cell(row=24, column=j, value=formula)
        c.number_format = "$#,##0"
        c.alignment = RIGHT

    # Annual payment
    ws.cell(row=25, column=1, value="Annual debt service (P&I)").font = BOLD
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=25, column=j, value=f"={col}24*12")
        c.number_format = "$#,##0"
        c.alignment = RIGHT

    # Add seller note monthly IO
    ws.cell(row=26, column=1, value="+ Seller note monthly IO").font = BOLD
    for j in range(2, 7):
        c = ws.cell(row=26, column=j, value=f"=$B$7*$B$8/12")
        c.number_format = "$#,##0"
        c.alignment = RIGHT

    # Total monthly payment
    ws.cell(row=27, column=1, value="TOTAL MONTHLY PAYMENT").font = Font(bold=True, color="0E3A66")
    ws.cell(row=27, column=1).fill = LIGHT_F
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=27, column=j, value=f"={col}24+{col}26")
        c.number_format = "$#,##0"
        c.alignment = RIGHT
        c.font = BOLD
        c.fill = LIGHT_F

    # Monthly cash flow
    ws.cell(row=28, column=1, value="Monthly NOI").font = BOLD
    for j in range(2, 7):
        c = ws.cell(row=28, column=j, value="=$B$17")
        c.number_format = "$#,##0"
        c.alignment = RIGHT

    ws.cell(row=29, column=1, value="MONTHLY CASH FLOW").font = Font(bold=True, color="065F46")
    ws.cell(row=29, column=1).fill = GREEN_F
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=29, column=j, value=f"={col}28-{col}27")
        c.number_format = "$#,##0"
        c.alignment = RIGHT
        c.font = BOLD
        c.fill = GREEN_F

    # DSCR
    ws.cell(row=30, column=1, value="DSCR (NOI / debt service)").font = BOLD
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=30, column=j, value=f"={col}28/{col}27")
        c.number_format = '0.00"x"'
        c.alignment = RIGHT

    # Annual cash flow
    ws.cell(row=31, column=1, value="ANNUAL CASH FLOW").font = Font(bold=True, color="065F46")
    ws.cell(row=31, column=1).fill = GREEN_F
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=31, column=j, value=f"={col}29*12")
        c.number_format = "$#,##0"
        c.alignment = RIGHT
        c.font = BOLD
        c.fill = GREEN_F

    # Notes
    ws.cell(row=33, column=1, value="WINNER: Freddie Mac SBL (column B)").font = Font(bold=True, color="065F46", size=12)
    ws.cell(row=33, column=1).fill = GREEN_F
    ws.merge_cells("A33:F33")

    ws.cell(row=34, column=1, value="Edit cells B3 (units), B4 (rent), B5 (vacancy), B6 (opex) to see live impact.")
    ws.cell(row=34, column=1).font = Font(italic=True, color="6B7280", size=9)
    ws.merge_cells("A34:F34")

    wb.save(path)


def send_email(paths, recipient):
    api_key = os.getenv("SENDGRID_API_KEY")
    from_email = os.getenv("SENDGRID_FROM_EMAIL", "info@rosstaxpreparation.com")
    html = """
    <div style="font-family:-apple-system,Segoe UI,sans-serif;max-width:680px;margin:0 auto;padding:24px;background:#f8fafc;">
      <div style="background:#fff;border-radius:12px;padding:28px;box-shadow:0 4px 12px rgba(0,0,0,0.06);">
        <h2 style="color:#0E3A66;margin:0 0 8px;">📊 Loan Comparison Analysis - Jasmine</h2>
        <p style="color:#6B7280;margin:0 0 18px;">5 lenders side-by-side - Monthly payment + cash flow + DSCR</p>

        <div style="background:#ECFDF5;border-left:4px solid #16A34A;padding:14px 18px;border-radius:8px;margin-bottom:18px;">
          <strong style="color:#065F46;">🏆 Ganador: Freddie Mac SBL</strong><br/>
          <span style="color:#065F46;font-size:13px;">$34,627/mo - 6.25% - 30 años - Non-recourse</span>
        </div>

        <h3 style="color:#0E5AA7;">📎 2 archivos adjuntos:</h3>
        <ol style="color:#1F2937;line-height:1.8;">
          <li><b>📄 Loan_Comparison_Analysis.pdf</b> - PDF con 5 lenders + 3 escenarios de renta</li>
          <li><b>📊 Loan_Comparison_LIVE.xlsx</b> - Excel con fórmulas LIVE - edita renta/opex y se recalcula</li>
        </ol>

        <p style="color:#6B7280;font-size:12px;margin-top:24px;border-top:1px solid #E5E7EB;padding-top:12px;">
          Ross House Rentals LLC
        </p>
      </div>
    </div>
    """
    message = Mail(from_email=from_email, to_emails=recipient,
                    subject="📊 Loan Comparison Analysis - Freddie/Fannie/CMBS/HUD/Bank (PDF + Excel)",
                    html_content=html)
    atts = []
    for p in paths:
        with open(p, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        ext = p.split(".")[-1].lower()
        mime = "application/pdf" if ext == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        atts.append(Attachment(FileContent(encoded), FileName(os.path.basename(p)),
                                FileType(mime), Disposition("attachment")))
    message.attachment = atts
    sg = SendGridAPIClient(api_key)
    resp = sg.send(message)
    print(f"SendGrid status: {resp.status_code}")


if __name__ == "__main__":
    p1 = "/tmp/Loan_Comparison_Analysis.pdf"
    p2 = "/tmp/Loan_Comparison_LIVE.xlsx"
    print("Building PDF...")
    build_pdf(p1)
    print(f"  -> {os.path.getsize(p1)} bytes")
    print("Building Excel...")
    build_excel(p2)
    print(f"  -> {os.path.getsize(p2)} bytes")
    print("Sending...")
    send_email([p1, p2], "yoandyross@gmail.com")
    print("Done!")
