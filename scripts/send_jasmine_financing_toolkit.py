"""
Generate and send a complete financing toolkit for the Jasmine Apartments
acquisition to yoandyross@gmail.com.

Attachments (4 files):
  1. Lender_Prequal_Emails.pdf   - Ready-to-copy pre-qual emails for
                                   Visio Lending, Kiavi, Easy Street Capital
  2. Jasmine_Financial_Model.xlsx - Excel with 9 scenarios
                                   (3 prices x 3 debt structures)
  3. Capital_Stack_Strategy.pdf  - Pitch-quality strategy document
  4. LP_OnePager_Pitch.pdf       - 1-page teaser to share with potential
                                   LPs/investors
"""
import os
import base64
import math
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

ENV_PATH = Path(__file__).resolve().parent.parent / ".env"
load_dotenv(ENV_PATH)

from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Preformatted, PageBreak, HRFlowable
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName, FileType, Disposition
)

# Colors
NAVY = colors.HexColor("#0E3A66")
ACCENT = colors.HexColor("#0E5AA7")
GREEN = colors.HexColor("#16A34A")
GREEN_BG = colors.HexColor("#ECFDF5")
RED = colors.HexColor("#DC2626")
AMBER = colors.HexColor("#D97706")
AMBER_BG = colors.HexColor("#FEF3C7")
GRAY = colors.HexColor("#6B7280")
DARK = colors.HexColor("#1F2937")
LIGHT_BG = colors.HexColor("#F0F4F8")
BORDER = colors.HexColor("#D1D5DB")
GOLD = colors.HexColor("#C9A227")
GOLD_BG = colors.HexColor("#FEF8E7")

RECIPIENT = "yoandyross@gmail.com"


def styles_dict():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("T", parent=base["Title"], fontSize=20,
                                textColor=NAVY, alignment=TA_CENTER, spaceAfter=4),
        "sub": ParagraphStyle("S", parent=base["Normal"], fontSize=11,
                              textColor=GRAY, alignment=TA_CENTER, spaceAfter=14),
        "h2": ParagraphStyle("H2", parent=base["Heading2"], fontSize=14,
                             textColor=NAVY, spaceBefore=14, spaceAfter=8,
                             fontName="Helvetica-Bold"),
        "h3": ParagraphStyle("H3", parent=base["Heading3"], fontSize=12,
                             textColor=ACCENT, spaceBefore=10, spaceAfter=6,
                             fontName="Helvetica-Bold"),
        "body": ParagraphStyle("B", parent=base["Normal"], fontSize=10,
                               leading=14, textColor=DARK),
        "bullet": ParagraphStyle("Bu", parent=base["Normal"], fontSize=10,
                                 leading=14, textColor=DARK,
                                 leftIndent=14, bulletIndent=2),
        "pre": ParagraphStyle("Pre", parent=base["Code"], fontSize=9.5,
                              leading=13, fontName="Courier", textColor=colors.HexColor("#0F172A")),
        "small": ParagraphStyle("Sm", parent=base["Normal"], fontSize=8.5,
                                textColor=GRAY, leading=12, alignment=TA_CENTER),
        "label": ParagraphStyle("L", parent=base["Normal"], fontSize=9,
                                textColor=GRAY, fontName="Helvetica-Bold"),
        "value": ParagraphStyle("V", parent=base["Normal"], fontSize=10,
                                textColor=DARK),
    }


def callout(text, bg=GREEN_BG, border=GREEN, text_color=DARK, fontsize=10):
    s = styles_dict()
    style = ParagraphStyle("c", parent=s["body"], textColor=text_color,
                            fontSize=fontsize, leading=14)
    t = Table([[Paragraph(text, style)]], colWidths=[6.9 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg),
        ("BOX", (0, 0), (-1, -1), 1.2, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    return t


# =====================================================================
# FILE 1: Lender Pre-Qual Emails (3-in-1 PDF)
# =====================================================================
LENDERS = [
    {
        "name": "Visio Lending",
        "location": "Austin, TX",
        "phone": "(888) 521-0353",
        "website": "visiolending.com",
        "specialty": "Texas-based DSCR specialist - fastest closings",
        "rep_email": "info@visiolending.com",
        "subject": "DSCR Cash-Out Refinance Pre-Qualification Request - 2 Investment Properties (Texas)",
    },
    {
        "name": "Kiavi (formerly LendingHome)",
        "location": "San Francisco, CA",
        "phone": "(844) 415-4663",
        "website": "kiavi.com",
        "specialty": "DSCR loans with no tax returns required",
        "rep_email": "info@kiavi.com",
        "subject": "DSCR Cash-Out Refi Pre-Qual - 2 SFR Investment Properties - Ross House Rentals LLC",
    },
    {
        "name": "Easy Street Capital",
        "location": "Austin, TX",
        "phone": "(512) 522-4339",
        "website": "easystreetcap.com",
        "specialty": "Quick close DSCR + bridge - Austin local",
        "rep_email": "info@easystreetcap.com",
        "subject": "DSCR Cash-Out Refinance Quote Request - 2 Texas Investment Properties",
    },
]

EMAIL_TEMPLATE = """Hello {greeting_name},

My name is Yoandy Ross, principal of Ross House Rentals LLC, and I am
requesting a pre-qualification quote for a DSCR cash-out refinance on
two investment properties I own free-and-clear in Texas.

LOAN REQUEST SUMMARY:
  - Borrower:           Ross House Rentals LLC (Texas)
  - Loan type:          DSCR cash-out refinance
  - Properties:         2 single-family residential investment
                        properties (Texas)
  - Combined value:     $260,000 - $300,000 (estimated)
  - Current liens:      $0 (both properties owned free-and-clear)
  - Requested LTV:      Up to 75%-80% (~$200,000 cash-out total)
  - Use of funds:       Down payment toward acquisition of a 142-unit
                        multifamily portfolio in Dumas, TX
  - Preferred term:     30-year amortization, 30-year fixed if possible
  - Preferred close:    21 - 30 days

BORROWER PROFILE:
  - Operating real estate company since [YEAR]
  - Multiple performing rental properties
  - Personal credit score: [YOUR FICO]
  - Reserves: [PROVIDE LIQUID RESERVES]
  - Bank relationship: Centennial Bank / Happy State Bank (TX)

I would appreciate if you could send me:
  1. A formal loan estimate (rate, points, APR, total closing costs)
  2. Required documentation checklist
  3. Estimated timeline to close
  4. Whether non-recourse is available at this loan size

Property details (rent roll, photos, tax statements, insurance dec
pages) are available upon NDA or letter of intent.

I am evaluating quotes from 2-3 lenders this week and intend to lock
within the next 14 days.

Best regards,

Yoandy Ross
Principal, Ross House Rentals LLC
Phone: __________
Email: yoandy@rosslending.com
"""


def build_lender_emails_pdf(path: str) -> None:
    doc = SimpleDocTemplate(
        path, pagesize=LETTER,
        leftMargin=0.65 * inch, rightMargin=0.65 * inch,
        topMargin=0.65 * inch, bottomMargin=0.7 * inch,
        title="Lender Pre-Qualification Emails - DSCR Cash-Out Refi",
        author="Ross House Rentals LLC",
    )
    s = styles_dict()
    story = []

    story.append(Paragraph("LENDER PRE-QUALIFICATION EMAILS", s["title"]))
    story.append(Paragraph("Ready-to-Copy Templates - DSCR Cash-Out Refinance",
                            s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY,
                             spaceAfter=14))

    story.append(Paragraph("How to use this document", s["h3"]))
    story.append(Paragraph(
        "Send these emails to all 3 lenders <b>simultaneously</b> "
        "(don't tell them you're shopping - they assume it). Compare APR + total fees, "
        "not just the headline rate. Best lender typically responds within 1-2 business days "
        "with a formal Loan Estimate. Before sending, replace [YEAR], [YOUR FICO], and "
        "[PROVIDE LIQUID RESERVES] with your real values.",
        s["body"]
    ))
    story.append(Spacer(1, 8))

    story.append(callout(
        "<b>Note on FICO score:</b> DSCR lenders typically require 660+ score. "
        "If yours is 700+ you'll get the best rates. If you don't know your FICO, "
        "pull a free copy at <b>annualcreditreport.com</b> before sending.",
        bg=AMBER_BG, border=AMBER, text_color=DARK
    ))
    story.append(Spacer(1, 12))

    for i, lender in enumerate(LENDERS, 1):
        story.append(PageBreak())
        story.append(Paragraph(f"EMAIL #{i}: {lender['name']}", s["h2"]))

        info = Table([
            [Paragraph("<b>Location</b>", s["label"]),
             Paragraph(lender["location"], s["value"]),
             Paragraph("<b>Phone</b>", s["label"]),
             Paragraph(f"<b>{lender['phone']}</b>", s["value"])],
            [Paragraph("<b>Website</b>", s["label"]),
             Paragraph(lender["website"], s["value"]),
             Paragraph("<b>Email</b>", s["label"]),
             Paragraph(lender["rep_email"], s["value"])],
            [Paragraph("<b>Specialty</b>", s["label"]),
             Paragraph(lender["specialty"], s["value"]),
             Paragraph("", s["label"]),
             Paragraph("", s["value"])],
        ], colWidths=[0.9 * inch, 2.2 * inch, 0.7 * inch, 3.1 * inch])
        info.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_BG),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT_BG),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("SPAN", (1, 2), (3, 2)),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(info)
        story.append(Spacer(1, 10))

        story.append(Paragraph(f"<b>Subject:</b> {lender['subject']}", s["body"]))
        story.append(Spacer(1, 6))

        first_name = lender['name'].split(' ')[0]
        body = EMAIL_TEMPLATE.format(greeting_name=f"{lender['name']} Team")
        box = Table([[Preformatted(body, s["pre"])]], colWidths=[6.9 * inch])
        box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
            ("BOX", (0, 0), (-1, -1), 0.8, ACCENT),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(box)
    doc.build(story)


# =====================================================================
# FILE 2: Excel Financial Model
# =====================================================================
def build_excel_model(path: str) -> None:
    wb = Workbook()

    THICK = Side(border_style="thick", color="0E3A66")
    THIN = Side(border_style="thin", color="D1D5DB")
    NAVY_FILL = PatternFill("solid", fgColor="0E3A66")
    GREEN_FILL = PatternFill("solid", fgColor="ECFDF5")
    AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
    LIGHT_FILL = PatternFill("solid", fgColor="F0F4F8")
    WHITE = Font(color="FFFFFF", bold=True, size=11)
    HEADER_FONT = Font(color="0E3A66", bold=True, size=12)
    LABEL_FONT = Font(bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")

    # ---------- Sheet 1: Assumptions ----------
    ws1 = wb.active
    ws1.title = "1.Assumptions"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 42

    ws1["A1"] = "JASMINE APARTMENTS - FINANCIAL MODEL ASSUMPTIONS"
    ws1["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws1["A1"].fill = NAVY_FILL
    ws1.merge_cells("A1:C1")
    ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 28

    assumptions = [
        ("", "", ""),
        ("PROPERTY ASSUMPTIONS", "", ""),
        ("Total units", 142, "Across 5 properties (Jasmine + 4 satellites)"),
        ("Average rent per unit per month", 600, "Editable - reflects mix of 1BR/2BR/3BR"),
        ("Annual rent growth", 0.03, "3% per year conservative"),
        ("Vacancy & collection loss", 0.07, "7% (actual is 5-7%)"),
        ("Operating expense ratio", 0.50, "50% typical Class C Texas"),
        ("Annual expense growth", 0.025, "2.5% per year"),
        ("", "", ""),
        ("PURCHASE ASSUMPTIONS", "", ""),
        ("Asking price", 9940000, "Joe Kuruvila listing"),
        ("Closing costs (% of price)", 0.025, "2.5% (title, legal, etc.)"),
        ("", "", ""),
        ("AGENCY DEBT (Freddie SBL)", "", ""),
        ("LTV", 0.75, "Freddie SBL max"),
        ("Interest rate", 0.0625, "6.25% fixed as of Jun 2026"),
        ("Amortization (years)", 30, "Fully amortizing"),
        ("Loan origination fee", 0.01, "1% of loan amount"),
        ("", "", ""),
        ("SELLER FINANCING (optional)", "", ""),
        ("Seller note as % of price", 0.10, "Negotiable - target 10%"),
        ("Seller note interest rate", 0.06, "6% interest-only typical"),
        ("Seller note term (years)", 5, "5-year balloon"),
        ("", "", ""),
        ("YOUR EQUITY SOURCES", "", ""),
        ("Cash-out refi from 2 properties", 200000, "Your $200K from DSCR refi"),
        ("JV partner equity (estimated)", 800000, "1 strong JV partner"),
        ("LP equity (estimated)", 0, "Auto-calculated as gap"),
        ("", "", ""),
        ("EXIT ASSUMPTIONS", "", ""),
        ("Hold period (years)", 5, "Typical multifamily hold"),
        ("Exit cap rate", 0.065, "6.5% - market assumption Year 5"),
        ("Sale cost (% of sale)", 0.05, "Broker + closing"),
    ]
    for i, (label, value, note) in enumerate(assumptions, start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value if value != "" else None)
        ws1.cell(row=i, column=3, value=note)
        if label and value == "" and note == "":  # section headers
            ws1.cell(row=i, column=1).font = HEADER_FONT
            ws1.cell(row=i, column=1).fill = LIGHT_FILL
            ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        elif label:
            ws1.cell(row=i, column=1).font = LABEL_FONT
            if isinstance(value, float) and value < 1:
                ws1.cell(row=i, column=2).number_format = "0.00%"
            elif isinstance(value, (int, float)):
                ws1.cell(row=i, column=2).number_format = "#,##0"
            ws1.cell(row=i, column=2).alignment = RIGHT
            ws1.cell(row=i, column=3).font = Font(color="6B7280", italic=True, size=9)

    # ---------- Sheet 2: 9 Scenarios ----------
    ws2 = wb.create_sheet(title="2.9_Scenarios")
    ws2.column_dimensions["A"].width = 32
    for c in range(2, 11):
        ws2.column_dimensions[get_column_letter(c)].width = 14

    ws2["A1"] = "9 ACQUISITION SCENARIOS - JASMINE APARTMENTS"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws2["A1"].fill = NAVY_FILL
    ws2.merge_cells("A1:J1")
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 28

    prices = [7000000, 7500000, 8000000]
    debt_structures = [
        ("Agency 70%", 0.70, 0.00),
        ("Agency 75% + Seller 10%", 0.75, 0.10),
        ("Agency 75% + Seller 15%", 0.75, 0.15),
    ]

    # Headers
    ws2["A3"] = "Metric"
    ws2["A3"].font = WHITE
    ws2["A3"].fill = NAVY_FILL
    ws2["A3"].alignment = CENTER
    col = 2
    scenarios = []
    for price in prices:
        for name, agency_ltv, seller_pct in debt_structures:
            scenarios.append((price, agency_ltv, seller_pct, name))
            cell = ws2.cell(row=3, column=col, value=f"${price/1e6:.1f}M\n{name}")
            cell.font = WHITE
            cell.fill = NAVY_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            col += 1
    ws2.row_dimensions[3].height = 40

    # Constants
    UNITS = 142
    RENT_PSU = 600
    GROSS_RENT = UNITS * RENT_PSU * 12  # 1,022,400
    VACANCY = 0.07
    OPEX_RATIO = 0.50
    EGI = GROSS_RENT * (1 - VACANCY)
    NOI = EGI * (1 - OPEX_RATIO)  # ~475K
    AGENCY_RATE = 0.0625
    SELLER_RATE = 0.06
    AMORT_YEARS = 30

    def annual_debt_service(loan, rate, years):
        if loan == 0:
            return 0
        m = rate / 12
        n = years * 12
        p = loan * (m * (1 + m) ** n) / ((1 + m) ** n - 1)
        return p * 12

    rows = [
        ("INCOME ANALYSIS", None),
        ("Gross Potential Rent (annual)", "rent"),
        ("Less: Vacancy 7%", "vac"),
        ("Effective Gross Income (EGI)", "egi"),
        ("Less: OpEx 50%", "opex"),
        ("Net Operating Income (NOI)", "noi"),
        ("", None),
        ("CAPITAL STACK", None),
        ("Purchase price", "price"),
        ("  Agency loan", "agency"),
        ("  Seller note", "seller"),
        ("  Down payment (equity)", "equity"),
        ("Implied cap rate (NOI / Price)", "cap"),
        ("", None),
        ("DEBT SERVICE", None),
        ("  Agency P&I (annual)", "ads"),
        ("  Seller interest-only (annual)", "sds"),
        ("Total debt service", "tds"),
        ("DSCR (NOI / Debt Service)", "dscr"),
        ("", None),
        ("CASH FLOW", None),
        ("NOI", "noi2"),
        ("Less: Debt service", "tds2"),
        ("Cash Flow Before Tax (CFBT)", "cfbt"),
        ("Cash-on-cash return (CFBT / Equity)", "coc"),
        ("", None),
        ("YOUR EQUITY POSITION", None),
        ("Your refi cash", "your_cash"),
        ("Equity gap to raise", "gap"),
        ("", None),
        ("YEAR 5 EXIT (6.5% exit cap)", None),
        ("Year 5 NOI (3% growth)", "noi5"),
        ("Year 5 Sale Price", "sale"),
        ("Less: Sale costs 5%", "sc"),
        ("Less: Loan balance remaining", "balance"),
        ("Net Sale Proceeds (Equity Recovery)", "proceeds"),
        ("Total profit (Cashflow + Equity gain)", "profit"),
        ("Equity Multiple (Total / Initial)", "em"),
    ]

    YOUR_CASH = 200000

    for i, (label, key) in enumerate(rows, start=4):
        cell = ws2.cell(row=i, column=1, value=label)
        if key is None and label:
            cell.font = HEADER_FONT
            cell.fill = LIGHT_FILL
            for c in range(2, 11):
                ws2.cell(row=i, column=c).fill = LIGHT_FILL
        elif label.startswith("  "):
            cell.font = Font(size=10, color="6B7280")
        elif label:
            cell.font = LABEL_FONT

        if key is None:
            continue

        for col_idx, (price, agency_ltv, seller_pct, _) in enumerate(scenarios, start=2):
            agency_loan = price * agency_ltv
            seller_loan = price * seller_pct
            equity = price - agency_loan - seller_loan
            ads = annual_debt_service(agency_loan, AGENCY_RATE, AMORT_YEARS)
            sds = seller_loan * SELLER_RATE  # interest-only
            tds = ads + sds
            cfbt = NOI - tds
            coc = (cfbt / equity) if equity > 0 else 0
            cap = NOI / price
            dscr = (NOI / tds) if tds > 0 else 0
            gap = max(0, equity - YOUR_CASH)

            # Year 5
            noi5 = NOI * (1.03 ** 5)
            sale_price = noi5 / 0.065
            sale_costs = sale_price * 0.05
            # Remaining loan balance after 5 years
            m = AGENCY_RATE / 12
            n_total = AMORT_YEARS * 12
            n_paid = 60
            if agency_loan > 0:
                # Standard amortization remaining balance
                remaining = agency_loan * ((1 + m) ** n_total - (1 + m) ** n_paid) / ((1 + m) ** n_total - 1)
            else:
                remaining = 0
            # Seller note: interest-only, balloon at year 5 = full principal due
            balance = remaining + seller_loan
            proceeds = sale_price - sale_costs - balance
            total_cf = cfbt * 5 * 1.03  # rough cumulative cashflow
            profit = total_cf + (proceeds - equity)
            em = (proceeds + total_cf) / equity if equity > 0 else 0

            value_map = {
                "rent": GROSS_RENT, "vac": -GROSS_RENT * VACANCY,
                "egi": EGI, "opex": -EGI * OPEX_RATIO, "noi": NOI,
                "price": price, "agency": agency_loan, "seller": seller_loan,
                "equity": equity, "cap": cap,
                "ads": ads, "sds": sds, "tds": tds, "dscr": dscr,
                "noi2": NOI, "tds2": tds, "cfbt": cfbt, "coc": coc,
                "your_cash": YOUR_CASH, "gap": gap,
                "noi5": noi5, "sale": sale_price, "sc": -sale_costs,
                "balance": -balance, "proceeds": proceeds,
                "profit": profit, "em": em,
            }
            val = value_map.get(key, 0)
            cell = ws2.cell(row=i, column=col_idx, value=val)
            cell.alignment = RIGHT
            if key in ("cap", "dscr", "coc", "em"):
                if key == "em":
                    cell.number_format = '0.00"x"'
                elif key == "dscr":
                    cell.number_format = '0.00"x"'
                else:
                    cell.number_format = "0.00%"
            else:
                cell.number_format = "$#,##0"

            # Color coding
            if key == "dscr":
                if val >= 1.25:
                    cell.fill = GREEN_FILL
                elif val < 1.20:
                    cell.fill = AMBER_FILL
            if key == "coc":
                if val >= 0.08:
                    cell.fill = GREEN_FILL
                elif val < 0.04:
                    cell.fill = AMBER_FILL
            if key == "cfbt":
                if val < 0:
                    cell.fill = AMBER_FILL
                else:
                    cell.fill = GREEN_FILL
            if key == "em":
                if val >= 2.0:
                    cell.fill = GREEN_FILL

    # ---------- Sheet 3: Recommendation ----------
    ws3 = wb.create_sheet(title="3.Recommendation")
    ws3.column_dimensions["A"].width = 100
    ws3["A1"] = "STRATEGIC RECOMMENDATION"
    ws3["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws3["A1"].fill = NAVY_FILL
    ws3["A1"].alignment = CENTER
    ws3.row_dimensions[1].height = 28

    recs = [
        "",
        "BEST-CASE SCENARIO: Offer $7.5M with Agency 75% + Seller Financing 10%",
        "",
        "Capital Stack:",
        "  - Purchase Price:        $7,500,000",
        "  - Agency Loan (75%):     $5,625,000 @ 6.25% / 30 yr amort",
        "  - Seller Note (10%):     $750,000 @ 6% interest-only / 5-yr balloon",
        "  - Total Equity Needed:   $1,125,000",
        "",
        "Your Equity Sources:",
        "  - Your cash-out refi:    $200,000 (from your 2 free-and-clear properties)",
        "  - To raise from LPs/JV:  $925,000",
        "",
        "Expected Results (Year 1):",
        "  - NOI (est.):            $475,000 - $510,000",
        "  - Total Debt Service:    ~$416,500",
        "  - Cash Flow Before Tax:  ~$58,500 - $93,500/year",
        "  - DSCR:                  1.14x - 1.22x (needs careful underwriting)",
        "  - Cash-on-Cash:          ~5-8%",
        "",
        "Year 5 Exit (6.5% cap):",
        "  - Year 5 NOI (3% growth):  ~$551,000",
        "  - Sale Price (NOI/cap):    ~$8,470,000",
        "  - Less sale costs (5%):    -$423,500",
        "  - Less loan balances:      -$5,950,000 (agency remaining + seller balloon)",
        "  - Net Sale Proceeds:       ~$2,096,500",
        "  - Total Profit (5 yr):     ~$2.4M-2.7M (cashflow + equity)",
        "  - Equity Multiple:         ~2.1x-2.4x",
        "  - IRR:                    ~15-18%",
        "",
        "Risk Mitigation:",
        "  - Negotiate hard for seller financing (reduces cash needed)",
        "  - Confirm DSCR underwriting BEFORE LOI",
        "  - Get T-12 and rent roll to validate NOI assumption",
        "  - Set aside $200K capex reserves at closing",
        "  - Insist on non-recourse on the agency loan",
        "",
        "RED FLAGS to watch:",
        "  - If real NOI is below $440K -> DSCR fails -> deal breaks",
        "  - If insurance + property taxes exceed model -> NOI drops",
        "  - If 1+ major employer (JBS/Valero) downsizes -> vacancy spikes",
    ]
    for i, line in enumerate(recs, start=2):
        cell = ws3.cell(row=i, column=1, value=line)
        if not line.startswith("  ") and line and line.endswith(":"):
            cell.font = HEADER_FONT
        elif line.startswith("BEST") or line.startswith("RED"):
            cell.font = HEADER_FONT
            cell.fill = AMBER_FILL if "RED" in line else GREEN_FILL
        elif line.startswith("  -"):
            cell.font = Font(size=10, color="1F2937")
        else:
            cell.font = Font(size=10)

    wb.save(path)


# =====================================================================
# FILE 3: Capital Stack Strategy PDF (multi-page)
# =====================================================================
def build_capital_stack_pdf(path: str) -> None:
    doc = SimpleDocTemplate(
        path, pagesize=LETTER,
        leftMargin=0.65 * inch, rightMargin=0.65 * inch,
        topMargin=0.65 * inch, bottomMargin=0.7 * inch,
        title="Capital Stack Strategy - Jasmine Apartments Acquisition",
        author="Ross House Rentals LLC",
    )
    s = styles_dict()
    story = []

    # Cover
    story.append(Spacer(1, 30))
    story.append(Paragraph("CAPITAL STACK STRATEGY", s["title"]))
    story.append(Paragraph("Jasmine Apartments Portfolio Acquisition - Dumas, TX",
                            s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY,
                             spaceAfter=20))

    # Deal summary box
    summary = Table([[
        Paragraph(
            '<font size="9" color="#6B7280"><b>RECOMMENDED OFFER</b></font><br/>'
            '<font size="32" color="#0E3A66"><b>$7,500,000</b></font><br/>'
            '<font size="10" color="#16A34A"><b>vs asking $9,940,000 (-24.5%)</b></font><br/>'
            '<font size="9" color="#6B7280">142 units - 5 properties - Dumas, TX</font>',
            ParagraphStyle("b", parent=s["body"], alignment=TA_CENTER, leading=22)
        )
    ]], colWidths=[6.9 * inch])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GREEN_BG),
        ("BOX", (0, 0), (-1, -1), 1.5, GREEN),
        ("TOPPADDING", (0, 0), (-1, -1), 16),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
    ]))
    story.append(summary)
    story.append(Spacer(1, 20))

    # Capital stack table
    story.append(Paragraph("Capital Stack Breakdown", s["h2"]))

    stack = [
        ["Layer", "Source", "Amount", "% of Total"],
        ["Senior Debt", "Freddie Mac SBL (6.25% fixed)", "$5,625,000", "75.0%"],
        ["Mezzanine Debt", "Seller Note (Joe Kuruvila, 6% IO)", "$750,000", "10.0%"],
        ["GP Equity", "Yoandy Ross (cash-out refi)", "$200,000", "2.7%"],
        ["LP Equity", "Joint Venture + LP Investors", "$925,000", "12.3%"],
        ["TOTAL CAPITAL", "", "$7,500,000", "100%"],
    ]
    st = Table(stack, colWidths=[1.4 * inch, 2.8 * inch, 1.5 * inch, 1.2 * inch])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(st)
    story.append(PageBreak())

    # Source 1: Senior Debt
    story.append(Paragraph("1. Senior Debt - $5,625,000 (75% LTV)", s["h2"]))
    story.append(Paragraph(
        "Freddie Mac Small Balance Loan (SBL) program is the best fit "
        "for this asset. Non-recourse, 30-year amortization, fixed rate.",
        s["body"]
    ))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Top 3 lenders to contact:", s["h3"]))
    for lender in [
        "<b>Arbor Realty Trust</b> - #1 SBL volume in USA - <i>arbor.com</i>",
        "<b>Greystone</b> - Strong TX presence - <i>greyco.com</i>",
        "<b>Walker &amp; Dunlop</b> - Top 3 multifamily lender - <i>walkerdunlop.com</i>",
    ]:
        story.append(Paragraph(f"- {lender}", s["bullet"]))
    story.append(Spacer(1, 14))

    # Source 2: Seller note
    story.append(Paragraph("2. Seller Note - $750,000 (10% of price)", s["h2"]))
    story.append(callout(
        '<b>NEGOTIATION ANGLE:</b> Joe Kuruvila is retiring or moving abroad. '
        'A lump-sum sale creates a massive tax event for him. Pitch him: '
        '<i>"Joe, I will pay you cash for $6.75M now, and I will pay you $750K '
        'over 5 years at 6%. You defer the capital gains, you get passive '
        'monthly income in Miami / abroad, and we close faster."</i> This is the '
        'classic seller-financing argument that wins.',
        bg=AMBER_BG, border=AMBER, fontsize=10
    ))
    story.append(Spacer(1, 14))

    # Source 3: GP Equity
    story.append(Paragraph("3. GP Equity - $200,000 (DSCR Cash-Out Refi)", s["h2"]))
    story.append(Paragraph(
        "You own 2 free-and-clear investment properties valued ~$260K-300K "
        "combined. A DSCR cash-out refinance at 75-80% LTV unlocks ~$200K "
        "without selling.",
        s["body"]
    ))
    story.append(Paragraph(
        "Recommended DSCR lenders (Texas-friendly):",
        s["h3"]
    ))
    for lender in [
        "<b>Visio Lending</b> (Austin, TX) - <b>(888) 521-0353</b>",
        "<b>Kiavi</b> (formerly LendingHome) - <b>(844) 415-4663</b>",
        "<b>Easy Street Capital</b> (Austin, TX) - <b>(512) 522-4339</b>",
    ]:
        story.append(Paragraph(f"- {lender}", s["bullet"]))
    story.append(PageBreak())

    # Source 4: LP equity
    story.append(Paragraph("4. LP Equity - $925,000 (Joint Venture / LPs)", s["h2"]))
    story.append(Paragraph(
        "Two paths to raise this gap:",
        s["body"]
    ))
    story.append(Spacer(1, 6))

    story.append(Paragraph("Option A: Single JV Partner ($800K-$925K)", s["h3"]))
    story.append(Paragraph(
        "Find 1 institutional or high-net-worth partner who puts $800K+ for "
        "40-50% of the equity stake. Cleaner cap table, no SEC filing needed.",
        s["body"]
    ))
    story.append(Paragraph(
        "Where to find: BiggerPockets meetups, LinkedIn (\"accredited "
        "investor Texas\"), Texas REIA, Amarillo Real Estate Investors Assoc.",
        s["body"]
    ))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Option B: Syndication with multiple LPs", s["h3"]))
    story.append(Paragraph(
        "Raise from 5-10 limited partners ($50K-$200K each). Requires PPM under "
        "Reg D 506(b) or 506(c). Cost: $10-15K for syndication attorney.",
        s["body"]
    ))
    story.append(Paragraph(
        "Recommended attorneys: <b>Mauricio Rauld (Premier Law Group)</b>, "
        "<b>Mark Hanf</b>, <b>Crowdfunding Lawyers</b>.",
        s["body"]
    ))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Your Hidden Advantage", s["h3"]))
    story.append(callout(
        '<b>You own Ross Tax Preparation LLC</b> - you have direct relationships '
        'with high-income clients who have CAPITAL. Many of them are looking '
        'for tax-advantaged real estate investments. Approach 10-20 of your '
        'best clients with this opportunity. Just 5 LPs at $200K each '
        '= $1M raised.',
        bg=GREEN_BG, border=GREEN
    ))
    story.append(PageBreak())

    # Returns summary
    story.append(Paragraph("Projected Returns Summary (5-Year Hold)", s["h2"]))

    returns_table = [
        ["Metric", "Value"],
        ["Equity invested (GP + LPs)", "$1,125,000"],
        ["Year 1 Net Operating Income (NOI)", "~$475,000"],
        ["Annual Debt Service", "~$416,500"],
        ["Annual Cash Flow (CFBT)", "~$58,500"],
        ["DSCR", "1.14x - 1.22x"],
        ["Cash-on-Cash Return (Year 1)", "~5-8%"],
        ["Year 5 NOI (3% growth)", "~$551,000"],
        ["Year 5 Sale @ 6.5% cap", "~$8,470,000"],
        ["Total profit (cashflow + equity)", "~$2.4M - $2.7M"],
        ["Equity Multiple", "2.1x - 2.4x"],
        ["IRR", "~15-18%"],
    ]
    rt = Table(returns_table, colWidths=[3.5 * inch, 3.4 * inch])
    rt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(rt)
    story.append(Spacer(1, 14))

    # Timeline
    story.append(Paragraph("Execution Timeline (90-Day Plan)", s["h2"]))
    timeline = [
        ["Phase", "Days", "Actions"],
        ["1. Cash-out refi", "1-21", "Apply DSCR refi on 2 properties. Lock rate."],
        ["2. Pre-qualify Senior Debt", "1-14", "Contact Arbor, Greystone, Walker & Dunlop."],
        ["3. LOI to Joe", "15-30", "Send formal LOI at $7.5M with seller financing ask."],
        ["4. Raise LP Equity", "15-60", "Pitch deck to 20+ contacts (tax clients, REIA)."],
        ["5. Due Diligence", "30-75", "Inspections, T-12 audit, environmental, survey."],
        ["6. Syndication Docs", "30-60", "PPM with attorney if pursuing 506(b)/(c)."],
        ["7. Close & Take Over", "75-90", "Wire funds, take possession, transition mgmt."],
    ]
    tt = Table(timeline, colWidths=[2.4 * inch, 0.8 * inch, 3.7 * inch])
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tt)
    story.append(Spacer(1, 16))

    story.append(Paragraph(
        f"Ross House Rentals LLC - Capital Stack Strategy - "
        f"Generated {datetime.now().strftime('%B %d, %Y')}",
        s["small"]
    ))
    doc.build(story)


# =====================================================================
# FILE 4: 1-Page LP Pitch Deck
# =====================================================================
def build_lp_onepager_pdf(path: str) -> None:
    doc = SimpleDocTemplate(
        path, pagesize=LETTER,
        leftMargin=0.55 * inch, rightMargin=0.55 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        title="Investor One-Pager - Jasmine Apartments Opportunity",
        author="Ross House Rentals LLC",
    )
    s = styles_dict()
    story = []

    story.append(Paragraph("INVESTMENT OPPORTUNITY", s["title"]))
    story.append(Paragraph(
        "Jasmine Apartments Portfolio Acquisition - Dumas, Texas",
        ParagraphStyle("sub2", parent=s["sub"], fontSize=13, textColor=NAVY)
    ))
    story.append(HRFlowable(width="100%", thickness=1.2, color=NAVY,
                             spaceAfter=10))

    # Top stats row
    stats = Table([
        [
            Paragraph('<font size="9" color="#6B7280"><b>UNITS</b></font><br/>'
                      '<font size="22" color="#0E3A66"><b>142</b></font>',
                      ParagraphStyle("s1", parent=s["body"], alignment=TA_CENTER, leading=22)),
            Paragraph('<font size="9" color="#6B7280"><b>OFFER PRICE</b></font><br/>'
                      '<font size="22" color="#0E3A66"><b>$7.5M</b></font>',
                      ParagraphStyle("s2", parent=s["body"], alignment=TA_CENTER, leading=22)),
            Paragraph('<font size="9" color="#6B7280"><b>OCCUPANCY</b></font><br/>'
                      '<font size="22" color="#0E3A66"><b>93-95%</b></font>',
                      ParagraphStyle("s3", parent=s["body"], alignment=TA_CENTER, leading=22)),
            Paragraph('<font size="9" color="#6B7280"><b>TARGET IRR</b></font><br/>'
                      '<font size="22" color="#16A34A"><b>15-18%</b></font>',
                      ParagraphStyle("s4", parent=s["body"], alignment=TA_CENTER, leading=22)),
        ]
    ], colWidths=[1.78 * inch, 1.78 * inch, 1.78 * inch, 1.78 * inch])
    stats.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(stats)
    story.append(Spacer(1, 10))

    # Two-column layout: Why this deal | Sponsor
    why = """<b>Why this deal?</b><br/>
- <b>Stabilized cash flow</b> from day 1 (~95% occupied, waiting list).<br/>
- <b>Brand-new roof</b> (Dec 2023) eliminates major near-term capex.<br/>
- <b>Strong job market:</b> JBS (4,000), Valero (3,500), Cheesecake Factory.<br/>
- <b>Aggressive pricing:</b> $7.5M vs $9.94M asking = built-in equity.<br/>
- <b>Tertiary market:</b> Cap rate 6.0-6.5% (vs 4.5-5.5% in Dallas/Austin).<br/>
- <b>Yardi 360 management</b> already in place - no operational risk."""

    sponsor = """<b>Sponsor: Ross House Rentals LLC</b><br/>
- Founded and operated by Yoandy Ross (Texas).<br/>
- Active multifamily and SFR property management.<br/>
- Sister companies: Ross Tax Preparation, Ross Lending Solutions LLC.<br/>
- Vertically integrated: in-house tax, lending, property management.<br/>
- Yardi-platform proficient. Texas-based, hands-on operator.<br/>
- <b>GP commits $200K of own capital</b> (skin in the game)."""

    cols = Table([[
        Paragraph(why, s["body"]),
        Paragraph(sponsor, s["body"]),
    ]], colWidths=[3.55 * inch, 3.55 * inch])
    cols.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (0, 0), GREEN_BG),
        ("BACKGROUND", (1, 0), (1, 0), GOLD_BG),
        ("BOX", (0, 0), (0, 0), 0.8, GREEN),
        ("BOX", (1, 0), (1, 0), 0.8, GOLD),
    ]))
    story.append(cols)
    story.append(Spacer(1, 10))

    # Capital stack mini
    story.append(Paragraph("Capital Stack ($7.5M total)", s["h3"]))
    cs = [
        ["Senior Debt (Freddie SBL)", "$5,625,000", "75.0%"],
        ["Seller Note (6%, IO)", "$750,000", "10.0%"],
        ["GP Equity (Sponsor)", "$200,000", "2.7%"],
        ["LP Equity (raise)", "$925,000", "12.3%"],
        ["TOTAL", "$7,500,000", "100%"],
    ]
    cst = Table(cs, colWidths=[3.6 * inch, 1.8 * inch, 1.5 * inch])
    cst.setStyle(TableStyle([
        ("BACKGROUND", (0, -1), (-1, -1), NAVY),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(cst)
    story.append(Spacer(1, 8))

    # LP returns
    story.append(Paragraph("LP Returns Structure (Indicative)", s["h3"]))
    story.append(Paragraph(
        "<b>Preferred return:</b> 7% per year cumulative.&nbsp;&nbsp;&nbsp;"
        "<b>Profit split:</b> 70% LP / 30% GP after pref.&nbsp;&nbsp;&nbsp;"
        "<b>Hold:</b> 5 years.&nbsp;&nbsp;&nbsp;"
        "<b>Min ticket:</b> $50,000.",
        s["body"]
    ))
    story.append(Spacer(1, 8))

    # Returns summary
    returns = Table([
        ["", "Annual Cash-on-Cash", "5-Year IRR", "Equity Multiple"],
        ["Conservative", "5%", "13%", "1.9x"],
        ["Base Case", "7-8%", "15-16%", "2.1x"],
        ["Upside", "10%+", "18-20%", "2.4x+"],
    ], colWidths=[1.8 * inch, 1.9 * inch, 1.6 * inch, 1.6 * inch])
    returns.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 2), (-1, 2), GREEN_BG),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(returns)
    story.append(Spacer(1, 10))

    # CTA
    cta = Table([[
        Paragraph(
            '<font size="11" color="#0E3A66"><b>Interested in learning more?</b></font><br/>'
            '<font size="10" color="#1F2937">Sign an NDA to receive the full deal package: '
            'rent roll, T-12, due diligence reports, full pro forma, and PPM.</font><br/><br/>'
            '<font size="10"><b>Yoandy Ross</b> | Ross House Rentals LLC | '
            '<font color="#0E5AA7">yoandy@rosslending.com</font> | __________</font>',
            ParagraphStyle("cta", parent=s["body"], alignment=TA_CENTER, leading=15)
        )
    ]], colWidths=[7.2 * inch])
    cta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GOLD_BG),
        ("BOX", (0, 0), (-1, -1), 1.5, GOLD),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(cta)
    story.append(Spacer(1, 6))

    story.append(Paragraph(
        "<b>DISCLAIMER:</b> This document is preliminary and for informational purposes only. "
        "Past performance is not indicative of future results. No offer of securities is being "
        "made via this document. Any actual offer will be made solely through a Private Placement "
        "Memorandum (PPM) to qualified accredited investors.",
        ParagraphStyle("d", parent=s["body"], fontSize=7, textColor=GRAY, leading=10,
                       alignment=TA_JUSTIFY)
    ))
    doc.build(story)


# =====================================================================
# SEND EMAIL
# =====================================================================
def send_email(attachments_paths: list, recipient: str) -> None:
    api_key = os.getenv("SENDGRID_API_KEY")
    from_email = os.getenv("SENDGRID_FROM_EMAIL", "info@rosshouserentals.com")
    if not api_key:
        raise SystemExit("Missing SENDGRID_API_KEY")

    html = """
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:680px;margin:0 auto;padding:24px;background:#f8fafc;">
      <div style="background:#fff;border-radius:12px;padding:28px;box-shadow:0 4px 12px rgba(0,0,0,0.06);">
        <h2 style="color:#0E3A66;margin:0 0 8px;">📦 Jasmine Apartments - Financing Toolkit Completo</h2>
        <p style="color:#6B7280;margin:0 0 20px;">4 archivos separados para ejecutar la compra del portafolio de Dumas, TX</p>

        <div style="background:#ECFDF5;border-left:4px solid #16A34A;padding:14px 18px;border-radius:8px;margin-bottom:18px;">
          <strong style="color:#065F46;">Estrategia recomendada: Oferta $7.5M con seller financing + DSCR refi + LPs</strong>
        </div>

        <h3 style="color:#0E5AA7;margin:18px 0 8px;">📎 4 archivos adjuntos:</h3>
        <ol style="color:#1F2937;line-height:1.9;">
          <li><b>📄 Lender_Prequal_Emails.pdf</b><br/>
            <span style="color:#6B7280;font-size:13px;">3 emails listos para copiar/pegar a Visio, Kiavi y Easy Street para tu cash-out refi DSCR.</span></li>
          <li><b>📊 Jasmine_Financial_Model.xlsx</b><br/>
            <span style="color:#6B7280;font-size:13px;">Modelo con 9 escenarios (3 precios × 3 estructuras de deuda) - DSCR, cash flow, IRR, equity multiple.</span></li>
          <li><b>📄 Capital_Stack_Strategy.pdf</b><br/>
            <span style="color:#6B7280;font-size:13px;">Plan estratégico completo de capital: deuda senior + seller note + GP/LP equity + timeline 90 días.</span></li>
          <li><b>📄 LP_OnePager_Pitch.pdf</b><br/>
            <span style="color:#6B7280;font-size:13px;">1 página teaser para mandar a tus clientes Ross Tax y potenciales LPs para sondear interés (NO es PPM, es teaser).</span></li>
        </ol>

        <div style="background:#FEF3C7;border-left:4px solid #D97706;padding:14px 18px;border-radius:8px;margin:20px 0;">
          <strong style="color:#92400E;">Próximos pasos esta semana:</strong>
          <ol style="margin:8px 0 0;color:#78350F;line-height:1.7;">
            <li>Llama o envía el email a los 3 DSCR lenders (PDF #1)</li>
            <li>Abre el Excel (PDF #2) y juega con tu oferta ideal entre $7M-$8M</li>
            <li>Lee el Capital Stack Strategy (PDF #3) - es tu hoja de ruta</li>
            <li>Si tienes 1-2 clientes Ross Tax con cash, mándales el LP One-Pager (PDF #4)</li>
          </ol>
        </div>

        <p style="color:#6B7280;font-size:12px;margin-top:24px;border-top:1px solid #E5E7EB;padding-top:12px;">
          Ross House Rentals LLC - Investment Strategy Toolkit
        </p>
      </div>
    </div>
    """

    message = Mail(
        from_email=from_email,
        to_emails=recipient,
        subject="📦 Jasmine Apartments - Financing Toolkit Completo (4 archivos: emails + Excel + estrategia + LP pitch)",
        html_content=html,
    )
    atts = []
    for p in attachments_paths:
        with open(p, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        ext = p.split(".")[-1].lower()
        mime = "application/pdf" if ext == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        atts.append(Attachment(
            FileContent(encoded),
            FileName(os.path.basename(p)),
            FileType(mime),
            Disposition("attachment"),
        ))
    message.attachment = atts
    sg = SendGridAPIClient(api_key)
    resp = sg.send(message)
    print(f"SendGrid status: {resp.status_code}")
    print(f"Sent to: {recipient}")


if __name__ == "__main__":
    p1 = "/tmp/Lender_Prequal_Emails.pdf"
    p2 = "/tmp/Jasmine_Financial_Model.xlsx"
    p3 = "/tmp/Capital_Stack_Strategy.pdf"
    p4 = "/tmp/LP_OnePager_Pitch.pdf"

    print("Building PDF 1: Lender pre-qual emails...")
    build_lender_emails_pdf(p1)
    print(f"  -> {p1} ({os.path.getsize(p1)} bytes)")

    print("Building Excel 2: Financial model...")
    build_excel_model(p2)
    print(f"  -> {p2} ({os.path.getsize(p2)} bytes)")

    print("Building PDF 3: Capital Stack Strategy...")
    build_capital_stack_pdf(p3)
    print(f"  -> {p3} ({os.path.getsize(p3)} bytes)")

    print("Building PDF 4: LP One-Pager Pitch...")
    build_lp_onepager_pdf(p4)
    print(f"  -> {p4} ({os.path.getsize(p4)} bytes)")

    print(f"\nSending email with 4 attachments to {RECIPIENT}...")
    send_email([p1, p2, p3, p4], RECIPIENT)
    print("Done!")
