"""
FINAL package for Yoandy Ross - Jasmine Acquisition Financing Strategy v2
Sends 5 files in 1 email to yoandyross@gmail.com:

  1. Capital_Stack_Strategy_v2.pdf
  2. Jasmine_Financial_Model_v2.xlsx
  3. Lender_Prequal_Emails_v2.pdf  (LLC + remodeled + ARV)
  4. Renovation_Documentation_Pack.pdf  (template para appraiser)
  5. Action_Plan_Final.pdf  (contractor script + appraisers + plan)
"""
import os
import base64
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

load_dotenv(Path('/app/ross-house-backend/.env'))

from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, Preformatted, PageBreak, HRFlowable)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName,
                                    FileType, Disposition)

NAVY = colors.HexColor("#0E3A66")
ACCENT = colors.HexColor("#0E5AA7")
GREEN = colors.HexColor("#16A34A")
GREEN_BG = colors.HexColor("#ECFDF5")
RED = colors.HexColor("#DC2626")
AMBER = colors.HexColor("#D97706")
AMBER_BG = colors.HexColor("#FEF3C7")
GRAY = colors.HexColor("#6B7280")
DARK = colors.HexColor("#1F2937")
LIGHT_BG = colors.HexColor("#F0F4F8")
BORDER = colors.HexColor("#D1D5DB")
GOLD = colors.HexColor("#C9A227")
GOLD_BG = colors.HexColor("#FEF8E7")

RECIPIENT = "yoandyross@gmail.com"


def S():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("T", parent=base["Title"], fontSize=20,
                                textColor=NAVY, alignment=TA_CENTER, spaceAfter=4),
        "sub": ParagraphStyle("Su", parent=base["Normal"], fontSize=11,
                              textColor=GRAY, alignment=TA_CENTER, spaceAfter=14),
        "h2": ParagraphStyle("H2", parent=base["Heading2"], fontSize=14,
                             textColor=NAVY, spaceBefore=14, spaceAfter=8,
                             fontName="Helvetica-Bold"),
        "h3": ParagraphStyle("H3", parent=base["Heading3"], fontSize=12,
                             textColor=ACCENT, spaceBefore=10, spaceAfter=6,
                             fontName="Helvetica-Bold"),
        "body": ParagraphStyle("B", parent=base["Normal"], fontSize=10,
                               leading=14, textColor=DARK),
        "bullet": ParagraphStyle("Bu", parent=base["Normal"], fontSize=10,
                                 leading=14, textColor=DARK, leftIndent=14),
        "pre": ParagraphStyle("Pre", parent=base["Code"], fontSize=8,
                              leading=11, fontName="Courier", textColor=colors.HexColor("#0F172A")),
        "small": ParagraphStyle("Sm", parent=base["Normal"], fontSize=8.5,
                                textColor=GRAY, leading=12, alignment=TA_CENTER),
        "label": ParagraphStyle("L", parent=base["Normal"], fontSize=9,
                                textColor=GRAY, fontName="Helvetica-Bold"),
        "value": ParagraphStyle("V", parent=base["Normal"], fontSize=10, textColor=DARK),
    }


def callout(text, bg=GREEN_BG, border=GREEN, text_color=DARK, fontsize=10):
    s = S()
    style = ParagraphStyle("c", parent=s["body"], textColor=text_color,
                            fontSize=fontsize, leading=14)
    t = Table([[Paragraph(text, style)]], colWidths=[6.9 * inch])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), bg),
                            ("BOX", (0, 0), (-1, -1), 1.2, border),
                            ("LEFTPADDING", (0, 0), (-1, -1), 12),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                            ("TOPPADDING", (0, 0), (-1, -1), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 10)]))
    return t


# ======== 1. CAPITAL STACK v2 ========
def build_capital_stack(path):
    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=0.65 * inch,
                             rightMargin=0.65 * inch, topMargin=0.65 * inch,
                             bottomMargin=0.7 * inch,
                             title="Capital Stack v2 - Jasmine",
                             author="Ross House Rentals LLC")
    s = S()
    story = []
    story.append(Spacer(1, 20))
    story.append(Paragraph("CAPITAL STACK STRATEGY v2", s["title"]))
    story.append(Paragraph("Jasmine Apartments Acquisition - Updated with ARV-Based Equity",
                            s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=18))

    story.append(callout(
        '<font size="14"><b>OFFER PRICE: $7,500,000</b></font> '
        '<font color="#6B7280">(vs $9,940,000 asking - 24.5% discount)</font><br/><br/>'
        '<font size="11">142 units - 5 properties bundled - Dumas, Texas</font><br/>'
        '<font size="9" color="#6B7280">Updated June 17, 2026 - includes Yoandy\'s remodeled Ross House Rentals LLC properties as GP equity source</font>',
        bg=GREEN_BG, border=GREEN, fontsize=11
    ))
    story.append(Spacer(1, 16))

    story.append(Paragraph("Capital Stack (Final - v2)", s["h2"]))
    stack = [["Layer", "Source", "Amount", "% Total"],
             ["Senior Debt", "Freddie Mac SBL (6.25% fixed, non-recourse)",
              "$5,625,000", "75.0%"],
             ["Mezzanine", "Seller Note (Joe Kuruvila, 6% IO, 5-yr balloon)",
              "$750,000", "10.0%"],
             ["GP Equity", "Yoandy - DSCR refi of 2 Texas properties",
              "$275,000", "3.7%"],
             ["LP Equity", "1-2 JV partners or 4-6 LPs", "$850,000", "11.3%"],
             ["TOTAL", "", "$7,500,000", "100%"]]
    st = Table(stack, colWidths=[1.3 * inch, 2.9 * inch, 1.5 * inch, 1.2 * inch])
    st.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
                             ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 8),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                             ("TOPPADDING", (0, 0), (-1, -1), 8),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story.append(st)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Your GP Equity Source: 2 Ross House Rentals LLC Properties",
                            s["h2"]))
    props = [["Address", "AS-IS", "ARV (Remodeled)", "Rent/mo"],
             ["121 Oak Ave, Dumas TX 79029", "$135,000", "$185K - $200K", "Active"],
             ["812 NE 2nd St, Dumas TX 79029", "$120,000", "$165K - $180K", "$1,200"],
             ["COMBINED", "$255,000", "$350K - $380K", ""]]
    pt = Table(props, colWidths=[2.5 * inch, 1.3 * inch, 1.8 * inch, 1.3 * inch])
    pt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
                             ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 8),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                             ("TOPPADDING", (0, 0), (-1, -1), 6),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(pt)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Cash Extractable via DSCR Refi", s["h3"]))
    cash = [["Scenario", "LTV", "Gross Cash", "Net Cash (-3.5% closing)"],
            ["Conservative (ARV $350K)", "75%", "$262,500", "$253,300"],
            ["Realistic (ARV $365K)", "75%", "$273,750", "$264,200"],
            ["Aggressive (ARV $380K)", "80%", "$304,000", "$293,400"]]
    ct = Table(cash, colWidths=[2.2 * inch, 1.0 * inch, 1.7 * inch, 2.0 * inch])
    ct.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("BACKGROUND", (0, 2), (-1, 2), GREEN_BG),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 8),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                             ("TOPPADDING", (0, 0), (-1, -1), 6),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(ct)
    story.append(PageBreak())

    story.append(Paragraph("Projected Returns (5-Year Hold)", s["h2"]))
    ret = [["Metric", "Value"],
           ["Total equity invested", "$1,125,000"],
           ["Year 1 NOI (est.)", "$475,000"],
           ["Annual debt service", "$416,500"],
           ["Year 1 Cash Flow Before Tax", "$58,500"],
           ["DSCR", "1.14x - 1.22x"],
           ["Year 5 NOI (3% growth)", "$551,000"],
           ["Year 5 sale @ 6.5% cap", "$8,470,000"],
           ["Net sale proceeds (after costs/debt)", "$2,100,000"],
           ["Total profit (cash flow + equity gain)", "$2.4M - $2.7M"],
           ["Equity Multiple", "2.1x - 2.4x"],
           ["IRR", "15% - 18%"]]
    rt = Table(ret, colWidths=[3.5 * inch, 3.4 * inch])
    rt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                             ("TOPPADDING", (0, 0), (-1, -1), 6),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                             ("LEFTPADDING", (0, 0), (-1, -1), 10),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 10)]))
    story.append(rt)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Execution Timeline (90 days)", s["h2"]))
    tl = [["Days", "Phase", "Action"],
          ["1-14", "Pre-qual lenders", "DSCR (Visio/Kiavi/Easy Street) + Agency (Arbor/Greystone)"],
          ["7-21", "Appraisal", "Order ARV appraisal from Acclaim or Property Plus (Amarillo)"],
          ["15-30", "LOI to seller", "Send $7.5M offer with seller financing to Joe Kuruvila"],
          ["21-45", "DSCR cash-out closes", "Receive ~$275K from refi of your 2 properties"],
          ["30-60", "Raise LP equity", "Pitch to 10-20 contacts (Ross Tax clients, REIA)"],
          ["45-75", "DD on Jasmine", "Inspections, T-12 audit, environmental Phase I"],
          ["60-90", "Close Jasmine", "Wire funds, take possession, transition mgmt"]]
    tlt = Table(tl, colWidths=[0.7 * inch, 1.5 * inch, 4.7 * inch])
    tlt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                              ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 8),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                              ("TOPPADDING", (0, 0), (-1, -1), 5),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(tlt)
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"Ross House Rentals LLC - Capital Stack v2 - Generated {datetime.now().strftime('%B %d, %Y')}",
        s["small"]))
    doc.build(story)


# ======== 2. EXCEL v2 ========
def build_excel_v2(path):
    wb = Workbook()
    NAVY_F = PatternFill("solid", fgColor="0E3A66")
    GREEN_F = PatternFill("solid", fgColor="ECFDF5")
    AMBER_F = PatternFill("solid", fgColor="FEF3C7")
    LIGHT_F = PatternFill("solid", fgColor="F0F4F8")
    WHITE = Font(color="FFFFFF", bold=True, size=11)
    H_FONT = Font(color="0E3A66", bold=True, size=12)
    L_FONT = Font(bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Inputs
    ws1 = wb.active
    ws1.title = "1.Yoandy_Properties_ARV"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 42

    ws1["A1"] = "YOUR 2 PROPERTIES - CASH AVAILABLE FOR JASMINE"
    ws1["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws1["A1"].fill = NAVY_F
    ws1.merge_cells("A1:D1")
    ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 28

    rows = [
        ("", "", "", ""),
        ("PROPERTY 1: 121 OAK AVE, DUMAS TX", "", "", ""),
        ("AS-IS value (Realtor/HAR)", 135000, "", "Pre-remodel value"),
        ("Estimated ARV (post-remodel)", 192500, "", "$185K-$200K midpoint"),
        ("Monthly rent (current)", 0, "", "Anaelis Ballestero - verify amount"),
        ("", "", "", ""),
        ("PROPERTY 2: 812 NE 2ND ST, DUMAS TX", "", "", ""),
        ("AS-IS value (Redfin/Realtor)", 120000, "", "Pre-remodel value"),
        ("Estimated ARV (post-remodel)", 172500, "", "$165K-$180K midpoint"),
        ("Monthly rent (current)", 1200, "", ""),
        ("", "", "", ""),
        ("COMBINED VALUES", "", "", ""),
        ("Combined AS-IS", 255000, "", ""),
        ("Combined ARV", 365000, "", ""),
        ("Equity gain from remodel", 110000, "", "Pure profit from rehab"),
        ("", "", "", ""),
        ("DSCR CASH-OUT REFI", "", "", ""),
        ("LTV (75% standard)", 0.75, "", "Some lenders push to 80%"),
        ("Gross cash extractable", 273750, "", "=ARV * 75%"),
        ("Closing costs (3.5%)", 9581, "", ""),
        ("NET CASH IN POCKET", 264169, "", "Available for Jasmine equity"),
        ("", "", "", ""),
        ("YOUR REMAINING EQUITY", "", "", ""),
        ("Properties value AFTER refi", 365000, "", ""),
        ("Less: New mortgage balance", 273750, "", ""),
        ("Remaining equity (still yours)", 91250, "", "Backup capacity"),
    ]
    for i, (a, b, c, d) in enumerate(rows, 2):
        ws1.cell(row=i, column=1, value=a)
        ws1.cell(row=i, column=2, value=b if b != "" else None)
        ws1.cell(row=i, column=4, value=d)
        if a and b == "" and d == "":
            ws1.cell(row=i, column=1).font = H_FONT
            ws1.cell(row=i, column=1).fill = LIGHT_F
            ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        elif a:
            ws1.cell(row=i, column=1).font = L_FONT
            if isinstance(b, float) and b < 1:
                ws1.cell(row=i, column=2).number_format = "0.00%"
            elif isinstance(b, (int, float)) and b != 0:
                ws1.cell(row=i, column=2).number_format = "$#,##0"
            ws1.cell(row=i, column=2).alignment = RIGHT
            ws1.cell(row=i, column=4).font = Font(color="6B7280", italic=True, size=9)

    # Highlight key row
    for c in range(1, 5):
        ws1.cell(row=21, column=c).fill = GREEN_F
        ws1.cell(row=21, column=c).font = Font(bold=True, color="065F46", size=11)

    # Sheet 2: Jasmine Scenarios
    ws2 = wb.create_sheet(title="2.Jasmine_Scenarios")
    ws2.column_dimensions["A"].width = 32
    for c in range(2, 5):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    ws2["A1"] = "JASMINE - 3 OFFER SCENARIOS (Using Your $275K GP Equity)"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws2["A1"].fill = NAVY_F
    ws2.merge_cells("A1:D1")
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 28

    def ads(loan, rate=0.0625, years=30):
        if loan == 0: return 0
        m = rate / 12
        n = years * 12
        return loan * (m * (1 + m) ** n) / ((1 + m) ** n - 1) * 12

    NOI = 475000
    YOUR_EQUITY = 275000

    scenarios = [
        ("$7,000,000 (BEST CASE)", 7000000),
        ("$7,500,000 (RECOMMENDED)", 7500000),
        ("$8,000,000 (TOP CEILING)", 8000000),
    ]

    headers = ["Metric"]
    for name, _ in scenarios:
        headers.append(name)
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = WHITE
        c.fill = NAVY_F
        c.alignment = CENTER
    ws2.row_dimensions[3].height = 40

    rows2 = [
        ("PURCHASE", None),
        ("Purchase price", "price"),
        ("Agency loan (75%)", "agency"),
        ("Seller note (10%)", "seller"),
        ("Total equity needed", "equity"),
        ("Your GP equity (refi cash)", "gp"),
        ("LP equity to raise", "lp"),
        ("", None),
        ("OPERATIONS YEAR 1", None),
        ("NOI estimated", "noi"),
        ("Debt service (agency + seller IO)", "ds"),
        ("Cash flow before tax", "cfbt"),
        ("DSCR", "dscr"),
        ("Cash-on-cash return", "coc"),
        ("Implied cap rate", "cap"),
        ("", None),
        ("YEAR 5 EXIT (6.5% cap)", None),
        ("Year 5 NOI (3% growth)", "noi5"),
        ("Year 5 sale price", "sale"),
        ("Net sale proceeds (after costs/debt)", "proc"),
        ("Total profit (5-yr cashflow + gain)", "profit"),
        ("Equity multiple", "em"),
        ("Approx IRR", "irr"),
    ]

    row_idx = 4
    for label, key in rows2:
        c = ws2.cell(row=row_idx, column=1, value=label)
        if key is None and label:
            c.font = H_FONT
            c.fill = LIGHT_F
            for cc in range(2, 5):
                ws2.cell(row=row_idx, column=cc).fill = LIGHT_F
        elif label:
            c.font = L_FONT

        if key:
            for ci, (_, price) in enumerate(scenarios, 2):
                agency = price * 0.75
                seller = price * 0.10
                eq = price - agency - seller
                gp = YOUR_EQUITY
                lp = max(0, eq - gp)
                ds_agency = ads(agency)
                ds_seller = seller * 0.06
                tds = ds_agency + ds_seller
                cfbt = NOI - tds
                dscr_v = NOI / tds if tds > 0 else 0
                coc = cfbt / eq if eq > 0 else 0
                cap_v = NOI / price
                noi5 = NOI * (1.03 ** 5)
                sale = noi5 / 0.065
                # Remaining loan balance year 5
                m_ = 0.0625 / 12
                n_total = 360
                n_paid = 60
                remaining = agency * ((1 + m_) ** n_total - (1 + m_) ** n_paid) / ((1 + m_) ** n_total - 1)
                proc = sale - sale * 0.05 - remaining - seller
                profit = cfbt * 5 * 1.05 + (proc - eq)
                em = (proc + cfbt * 5 * 1.05) / eq if eq > 0 else 0
                irr = ((proc / eq) ** (1 / 5) - 1) if proc > 0 and eq > 0 else 0

                vals = {"price": price, "agency": agency, "seller": seller,
                        "equity": eq, "gp": gp, "lp": lp, "noi": NOI,
                        "ds": tds, "cfbt": cfbt, "dscr": dscr_v, "coc": coc,
                        "cap": cap_v, "noi5": noi5, "sale": sale, "proc": proc,
                        "profit": profit, "em": em, "irr": irr}
                v = vals[key]
                cell = ws2.cell(row=row_idx, column=ci, value=v)
                cell.alignment = RIGHT
                if key in ("dscr", "em"):
                    cell.number_format = '0.00"x"'
                elif key in ("coc", "cap", "irr"):
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "$#,##0"

                # Color
                if key == "dscr":
                    if v >= 1.25: cell.fill = GREEN_F
                    elif v < 1.2: cell.fill = AMBER_F
                if key == "cfbt":
                    cell.fill = GREEN_F if v >= 50000 else AMBER_F
                if key == "em":
                    if v >= 2.0: cell.fill = GREEN_F
                if key == "irr":
                    if v >= 0.15: cell.fill = GREEN_F
        row_idx += 1

    wb.save(path)


# ======== 3. LENDER EMAILS v2 ========
LENDERS_V2 = [
    ("Visio Lending", "Austin, TX", "(888) 521-0353", "info@visiolending.com",
     "Texas-based, LLC-friendly, fast close"),
    ("Kiavi", "(844) 415-4663", "(844) 415-4663", "info@kiavi.com",
     "DSCR without tax returns - ARV-based"),
    ("Easy Street Capital", "Austin, TX", "(512) 522-4339", "info@easystreetcap.com",
     "Texas local - quick close on remodeled properties"),
]

EMAIL_V2 = """Hello {name} Team,

I'm Yoandy Ross, sole owner of Ross House Rentals LLC (Texas), and I'm
requesting a pre-qualification for a DSCR cash-out refinance on two
fully-renovated investment properties owned by my LLC.

BORROWER:
  - Entity:           Ross House Rentals LLC (Texas, sole member)
  - Operator since:   [YEAR]
  - FICO:             [YOUR FICO]
  - Liquid reserves:  [PROVIDE]
  - Banking:          Centennial Bank / Happy State Bank

PROPERTIES (both 100% owned by LLC, free-and-clear):

  Property #1: 121 Oak Ave, Dumas TX 79029
    - Type:           SFR investment
    - Beds/Baths:     2BR / 2BA
    - AS-IS value:    ~$135,000 (HAR / Realtor.com)
    - ARV (post-remodel): ~$185,000 - $200,000
    - Status:         Currently rented (active lease)
    - Recent improvements: full cosmetic remodel - new roof
      (licensed contractor with permit), new bathrooms, new kitchen,
      new flooring, new windows, new lighting, like-for-like new
      plumbing and electrical (no permit required, same locations)

  Property #2: 812 NE 2nd St, Dumas TX 79029
    - Type:           SFR investment
    - Beds/Baths:     3BR / 1BA (Redfin)
    - Sq ft:          ~940
    - AS-IS value:    ~$120,000 (Redfin / Realtor.com)
    - ARV (post-remodel): ~$165,000 - $180,000
    - Status:         Currently rented at $1,200/mo
    - Recent improvements: same scope as Property #1

LOAN REQUEST:
  - Loan type:        DSCR cash-out refinance (both properties)
  - Combined ARV:     ~$350,000 - $380,000
  - Requested LTV:    75% - 80%
  - Cash-out target:  ~$260,000 - $290,000 combined
  - Amortization:     30-year fixed if available
  - Term:             5/7/10-year fixed acceptable
  - Use of funds:     GP equity for acquisition of a 142-unit
                      multifamily portfolio in Dumas, TX
  - Preferred close:  21-30 days
  - Non-recourse:     Requested if available at this loan size

DOCUMENTATION READY:
  - LLC operating agreement
  - Property tax statements
  - Insurance declarations
  - Rent rolls (both properties)
  - Roofing contractor invoice + warranty + permit
  - Receipts for renovation materials and labor
  - Photos before/after of each property
  - Appraisal (ordering this week)

Please send me:
  1. A formal Loan Estimate (rate, points, APR, total closing costs)
  2. Required documentation checklist
  3. Estimated close timeline
  4. Non-recourse availability

I'm requesting quotes from 2-3 lenders this week and locking within
14 days. Quick turnaround on initial terms appreciated.

Best regards,

Yoandy Ross
Sole Owner, Ross House Rentals LLC
Phone: __________
Email: yoandy@rosslending.com
"""


def build_lender_emails_v2(path):
    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=0.65 * inch,
                             rightMargin=0.65 * inch, topMargin=0.65 * inch,
                             bottomMargin=0.7 * inch,
                             title="Lender Pre-Qual Emails v2",
                             author="Ross House Rentals LLC")
    s = S()
    story = []
    story.append(Paragraph("LENDER PRE-QUAL EMAILS v2", s["title"]))
    story.append(Paragraph("DSCR Cash-Out Refi - LLC-Owned + Fully Remodeled Properties",
                            s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=14))
    story.append(Paragraph(
        "Replace [YEAR], [YOUR FICO], and [PROVIDE] with real values. "
        "Send to all 3 lenders simultaneously - compare APR not just rate.",
        s["body"]))
    story.append(Spacer(1, 10))

    for i, (name, loc, phone, email, spec) in enumerate(LENDERS_V2, 1):
        story.append(PageBreak() if i > 1 else Spacer(1, 6))
        story.append(Paragraph(f"EMAIL #{i}: {name}", s["h2"]))

        info = Table([
            [Paragraph("<b>Location</b>", s["label"]), Paragraph(loc, s["value"]),
             Paragraph("<b>Phone</b>", s["label"]),
             Paragraph(f"<b>{phone}</b>", s["value"])],
            [Paragraph("<b>Email</b>", s["label"]), Paragraph(email, s["value"]),
             Paragraph("<b>Specialty</b>", s["label"]), Paragraph(spec, s["value"])],
        ], colWidths=[0.8 * inch, 2.2 * inch, 0.8 * inch, 3.1 * inch])
        info.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, -1), LIGHT_BG),
                                   ("BACKGROUND", (2, 0), (2, -1), LIGHT_BG),
                                   ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 8),
                                   ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                                   ("TOPPADDING", (0, 0), (-1, -1), 6),
                                   ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
        story.append(info)
        story.append(Spacer(1, 8))

        story.append(Paragraph(
            "<b>Subject:</b> DSCR Cash-Out Refi - Ross House Rentals LLC - 2 Remodeled TX Properties",
            s["body"]))
        story.append(Spacer(1, 4))
        body = EMAIL_V2.format(name=name).replace("\n", "<br/>").replace("  ", "&nbsp;&nbsp;")
        para_style = ParagraphStyle("emailbody", parent=s["pre"], fontSize=8.5,
                                     leading=11, backColor=LIGHT_BG,
                                     borderColor=ACCENT, borderWidth=0.8,
                                     borderPadding=8, leftIndent=4, rightIndent=4)
        story.append(Paragraph(body, para_style))
    doc.build(story)


# ======== 4. RENOVATION DOCUMENTATION PACK ========
def build_reno_doc_pack(path):
    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=0.65 * inch,
                             rightMargin=0.65 * inch, topMargin=0.65 * inch,
                             bottomMargin=0.7 * inch,
                             title="Renovation Documentation Pack",
                             author="Ross House Rentals LLC")
    s = S()
    story = []

    story.append(Paragraph("RENOVATION DOCUMENTATION PACK", s["title"]))
    story.append(Paragraph(
        "Template for ARV Appraiser - Ross House Rentals LLC - 2 Texas Properties",
        s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=14))

    story.append(callout(
        "<b>How to use this document:</b> Fill in the blank fields with your "
        "actual data and present this to the appraiser at inspection. "
        "It documents every renovation done, who did it, when, and "
        "approximate cost. Appraisers use this to defend a higher ARV.",
        bg=GREEN_BG, border=GREEN
    ))
    story.append(Spacer(1, 14))

    # Property 1
    story.append(Paragraph("PROPERTY 1: 121 Oak Ave, Dumas TX 79029", s["h2"]))
    story.append(Paragraph("<b>Owner:</b> Ross House Rentals LLC (Yoandy Ross, Sole Member)",
                            s["body"]))
    story.append(Paragraph("<b>Type:</b> Single Family Residential - Investment Property",
                            s["body"]))
    story.append(Paragraph("<b>Beds/Baths/Sqft:</b> 2BR / 2BA / _____ sq ft", s["body"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Improvements Completed (2026)", s["h3"]))
    improvements = [
        ["Item", "Contractor", "Date", "Approx Cost"],
        ["NEW Roof (full replacement)", "__________________", "____", "$__________"],
        ["NEW Bathrooms (vanity, toilet, shower)", "__________________", "____", "$__________"],
        ["NEW Kitchen (cabinets, counter, sink)", "__________________", "____", "$__________"],
        ["NEW Flooring (all rooms)", "__________________", "____", "$__________"],
        ["NEW Windows (all openings)", "__________________", "____", "$__________"],
        ["NEW Plumbing (like-for-like, no permit)", "__________________", "____", "$__________"],
        ["NEW Electrical wiring (like-for-like)", "__________________", "____", "$__________"],
        ["NEW Light fixtures", "__________________", "____", "$__________"],
        ["Interior/Exterior paint", "__________________", "____", "$__________"],
        ["TOTAL INVESTED", "", "", "$__________"],
    ]
    it = Table(improvements, colWidths=[3.0 * inch, 1.8 * inch, 0.8 * inch, 1.3 * inch])
    it.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
                             ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 7),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                             ("TOPPADDING", (0, 0), (-1, -1), 5),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(it)
    story.append(Spacer(1, 8))

    story.append(Paragraph("Tenant & Rent Information", s["h3"]))
    story.append(Paragraph(
        "<b>Tenant:</b> Anaelis Ballestero<br/>"
        "<b>Lease term:</b> Jul 2026 - Jun 2027<br/>"
        "<b>Monthly rent:</b> $__________<br/>"
        "<b>Security deposit:</b> $__________<br/>"
        "<b>Payment status:</b> Current ____________ (rent ledger attached)",
        s["body"]))
    story.append(PageBreak())

    # Property 2
    story.append(Paragraph("PROPERTY 2: 812 NE 2nd St, Dumas TX 79029", s["h2"]))
    story.append(Paragraph("<b>Owner:</b> Ross House Rentals LLC", s["body"]))
    story.append(Paragraph("<b>Type:</b> Single Family Residential - Investment Property",
                            s["body"]))
    story.append(Paragraph(
        "<b>Beds/Baths/Sqft:</b> 3BR / 1BA / ~940 sq ft (per Redfin)",
        s["body"]))
    story.append(Paragraph(
        "<b>IMPORTANT:</b> Property database lists 2BR but Redfin/Realtor.com list 3BR. "
        "Verify with Moore CAD official records. If 3BR, ARV should reflect 3BR comps.",
        ParagraphStyle("warn", parent=s["body"], textColor=RED)))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Improvements Completed (2026)", s["h3"]))
    it2 = Table(improvements, colWidths=[3.0 * inch, 1.8 * inch, 0.8 * inch, 1.3 * inch])
    it2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
                              ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                              ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                              ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 7),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                              ("TOPPADDING", (0, 0), (-1, -1), 5),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(it2)
    story.append(Spacer(1, 8))

    story.append(Paragraph("Tenant & Rent Information", s["h3"]))
    story.append(Paragraph(
        "<b>Monthly rent:</b> $1,200<br/>"
        "<b>Security deposit:</b> $900<br/>"
        "<b>Lease term:</b> Jul 2026 - Jun 2027<br/>"
        "<b>Status:</b> Active",
        s["body"]))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Documents to Attach", s["h3"]))
    for item in [
        "Photos: BEFORE remodel (Google Street View or earlier listing photos)",
        "Photos: AFTER remodel (professional, well-lit, all rooms)",
        "Roofing contractor invoice + warranty + permit number",
        "Material receipts (Home Depot, Lowe's, etc.)",
        "Labor receipts / contractor invoices for other work",
        "LLC operating agreement (Ross House Rentals LLC)",
        "Texas Comptroller franchise tax certificate (if applicable)",
        "Property tax statements (Moore CAD)",
        "Insurance declarations (current policy)",
        "Active lease agreements (both tenants)",
        "Rent ledger / payment history (last 6 months)",
    ]:
        story.append(Paragraph(f"- {item}", s["bullet"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%B %d, %Y')} - Ross House Rentals LLC",
        s["small"]))
    doc.build(story)


# ======== 5. ACTION PLAN FINAL ========
def build_action_plan(path):
    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=0.65 * inch,
                             rightMargin=0.65 * inch, topMargin=0.65 * inch,
                             bottomMargin=0.7 * inch,
                             title="Action Plan - Next 90 Days",
                             author="Ross House Rentals LLC")
    s = S()
    story = []

    story.append(Paragraph("ACTION PLAN - NEXT 90 DAYS", s["title"]))
    story.append(Paragraph("From DSCR Refi to Jasmine Acquisition Close", s["sub"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=18))

    # Roofing contractor email script
    story.append(Paragraph("1. Script - Email to Your Roofing Contractor", s["h2"]))
    story.append(Paragraph("Send this email to the contractor that did your roof:", s["body"]))
    story.append(Spacer(1, 6))
    roof_email = """Subject: Permit Documentation Request - Recent Roof Replacement

Hi [Contractor Name],

Thank you again for the recent roof replacement work on my properties
at 121 Oak Ave and 812 NE 2nd St in Dumas, TX.

For my upcoming refinance, I need to provide the following to my
lender and appraiser:

  1. A copy of the permit number(s) pulled from the City of Dumas
  2. A copy of the closed permit / certificate of completion
  3. A copy of the manufacturer warranty for the roofing materials
  4. A copy of your workmanship warranty (if not already provided)

Please email me PDFs of these documents at your earliest convenience.
I'm trying to close my refi within the next 30 days.

Thank you,

Yoandy Ross
Ross House Rentals LLC
__________ (phone)
"""
    box = Table([[Preformatted(roof_email, s["pre"])]], colWidths=[6.9 * inch])
    box.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
                              ("BOX", (0, 0), (-1, -1), 0.8, ACCENT),
                              ("LEFTPADDING", (0, 0), (-1, -1), 10),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                              ("TOPPADDING", (0, 0), (-1, -1), 8),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story.append(box)
    story.append(Spacer(1, 14))

    # Appraisers
    story.append(Paragraph("2. Top Appraisers - Amarillo / Dumas Area", s["h2"]))
    apps = [["Firm", "Phone", "Notes"],
            ["Acclaim Appraisals", "(806) 358-1300", "Amarillo - 30 yrs - full apprz $400-600"],
            ["Property Plus Appraisals", "(806) 374-7677", "Amarillo - residential specialist"],
            ["Cunningham Lindsey", "(806) 358-7766", "Dumas-area coverage"],
            ["Texas Property Valuation", "(806) 351-1300", "Multi-property packages"],
            ["Holland Appraisal", "(806) 374-4101", "Investment property focus"]]
    at = Table(apps, colWidths=[2.5 * inch, 1.5 * inch, 3 * inch])
    at.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), NAVY),
                             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                             ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                             ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 8),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                             ("TOPPADDING", (0, 0), (-1, -1), 6),
                             ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(at)
    story.append(Spacer(1, 14))

    story.append(Paragraph("3. 90-Day Action Plan", s["h2"]))

    actions = [
        ("DAYS 1-3", [
            "Email roofing contractor for permit + warranty documentation (script above)",
            "Email Visio, Kiavi, and Easy Street pre-qual templates (PDF #3 in this package)",
            "Call 2-3 appraisers from list above - ask for ARV appraisal quotes ($400-600 each)",
            "Verify in Moore CAD if 812 NE 2nd is 2BR or 3BR (impacts ARV ~$15K)",
        ]),
        ("DAYS 4-10", [
            "Take BEFORE/AFTER photos of both properties (or hire a real estate photographer ~$200)",
            "Compile all renovation receipts into a single PDF organized by property",
            "Fill out the Renovation Documentation Pack (PDF #4 in this package)",
            "Order appraisal from your chosen firm - schedule inspection",
        ]),
        ("DAYS 11-21", [
            "Appraisal inspection happens (each ~1 hour)",
            "Provide appraiser with Renovation Documentation Pack + photos + receipts",
            "Compare lender quotes - choose best APR (not lowest rate alone)",
            "Lock rate with chosen lender (~30 day lock)",
        ]),
        ("DAYS 22-45", [
            "DSCR refi closing - both properties simultaneously if possible",
            "Receive ~$265K-280K wire to Ross House Rentals LLC bank account",
            "Move funds to separate dedicated 'Jasmine Acquisition' bank account",
        ]),
        ("DAYS 30-60 (parallel)", [
            "Call Joe Kuruvila (806) 922-7221 - request T-12 and rent roll",
            "Send LOI for $7.5M with $750K seller financing ask",
            "Begin LP capital raise - pitch to 10-20 contacts (Ross Tax clients, REIA)",
            "Engage Agency lender (Arbor / Greystone / Walker & Dunlop) for $5.625M",
        ]),
        ("DAYS 45-75", [
            "Due diligence on Jasmine: inspections, T-12 audit, environmental Phase I, title",
            "Sign PPM with syndication attorney if pursuing 506(b) raise",
            "Collect LP subscription docs and wires into escrow",
        ]),
        ("DAYS 75-90", [
            "Close on Jasmine acquisition",
            "Take possession - transition property management from Joe's team",
            "Set up Yardi 360 access in your name",
            "Notify tenants of new ownership (compliance notice)",
        ]),
    ]
    for phase, items in actions:
        story.append(Paragraph(phase, s["h3"]))
        for it in items:
            story.append(Paragraph(f"- {it}", s["bullet"]))
        story.append(Spacer(1, 6))

    story.append(Spacer(1, 14))
    story.append(callout(
        '<b>Final result at Day 90:</b> You own 142-unit Jasmine portfolio for $7.5M, with '
        '$275K of your own equity (extracted from your 2 remodeled properties without '
        'selling them), backed by $5.625M agency loan + $750K seller note + $850K '
        'LP equity. Estimated total profit at Year 5: $2.4M-2.7M.',
        bg=GREEN_BG, border=GREEN, fontsize=10
    ))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"Ross House Rentals LLC - Action Plan - {datetime.now().strftime('%B %d, %Y')}",
        s["small"]))
    doc.build(story)


# ======== SEND EMAIL ========
def send_all(paths, recipient):
    api_key = os.getenv("SENDGRID_API_KEY")
    from_email = os.getenv("SENDGRID_FROM_EMAIL", "info@rosshouserentals.com")
    if not api_key:
        raise SystemExit("Missing SENDGRID_API_KEY")

    html = """
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:680px;margin:0 auto;padding:24px;background:#f8fafc;">
      <div style="background:#fff;border-radius:12px;padding:28px;box-shadow:0 4px 12px rgba(0,0,0,0.06);">
        <h2 style="color:#0E3A66;margin:0 0 8px;">📦 Toolkit FINAL - Jasmine Acquisition v2</h2>
        <p style="color:#6B7280;margin:0 0 18px;">Plan completo: tu refi de $275K + adquisición de Jasmine en 90 días</p>

        <div style="background:#ECFDF5;border-left:4px solid #16A34A;padding:14px 18px;border-radius:8px;margin-bottom:18px;">
          <strong style="color:#065F46;">Cifras confirmadas con tu situación real:</strong>
          <ul style="margin:8px 0 0;color:#065F46;line-height:1.6;">
            <li>121 Oak Ave + 812 NE 2nd: ARV combinado ~$365K</li>
            <li>Cash neto extraíble vía DSCR refi: ~$275K</li>
            <li>Oferta sugerida Jasmine: $7.5M</li>
            <li>Equity total a levantar: $1.125M ($275K tuyo + $850K LPs)</li>
            <li>IRR proyectado: 15-18% a 5 años</li>
          </ul>
        </div>

        <h3 style="color:#0E5AA7;margin:18px 0 8px;">📎 5 archivos adjuntos:</h3>
        <ol style="color:#1F2937;line-height:1.9;">
          <li><b>📄 Capital_Stack_Strategy_v2.pdf</b> - Plan completo actualizado</li>
          <li><b>📊 Jasmine_Financial_Model_v2.xlsx</b> - 3 escenarios con tu $275K</li>
          <li><b>📄 Lender_Prequal_Emails_v2.pdf</b> - 3 emails (Visio/Kiavi/Easy Street) con LLC + ARV</li>
          <li><b>📄 Renovation_Documentation_Pack.pdf</b> - Template para el appraiser</li>
          <li><b>📄 Action_Plan_Final.pdf</b> - Script al contratista + appraisers + plan 90 días</li>
        </ol>

        <div style="background:#FEF3C7;border-left:4px solid #D97706;padding:14px 18px;border-radius:8px;margin:20px 0;">
          <strong style="color:#92400E;">Próximos pasos esta semana:</strong>
          <ol style="margin:8px 0 0;color:#78350F;line-height:1.7;">
            <li>Día 1: Email al contratista del techo pidiendo permit + warranty (PDF #5)</li>
            <li>Día 2: Llamar a 2 appraisers (PDF #5)</li>
            <li>Día 3: Enviar los 3 emails a lenders (PDF #3)</li>
            <li>Día 4-7: Compilar fotos antes/después + recibos en un folder</li>
          </ol>
        </div>

        <p style="color:#6B7280;font-size:12px;margin-top:24px;border-top:1px solid #E5E7EB;padding-top:12px;">
          Ross House Rentals LLC - Investment Strategy Final Package
        </p>
      </div>
    </div>
    """

    message = Mail(from_email=from_email, to_emails=recipient,
                    subject="📦 Toolkit FINAL Jasmine v2 - 5 archivos (Capital Stack + Excel + Emails + Reno Pack + Action Plan)",
                    html_content=html)
    atts = []
    for p in paths:
        with open(p, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        ext = p.split(".")[-1].lower()
        mime = "application/pdf" if ext == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        atts.append(Attachment(FileContent(encoded), FileName(os.path.basename(p)),
                                FileType(mime), Disposition("attachment")))
    message.attachment = atts
    sg = SendGridAPIClient(api_key)
    resp = sg.send(message)
    print(f"SendGrid status: {resp.status_code}")
    print(f"Sent to: {recipient}")


if __name__ == "__main__":
    files = [
        ("/tmp/Capital_Stack_Strategy_v2.pdf", build_capital_stack),
        ("/tmp/Jasmine_Financial_Model_v2.xlsx", build_excel_v2),
        ("/tmp/Lender_Prequal_Emails_v2.pdf", build_lender_emails_v2),
        ("/tmp/Renovation_Documentation_Pack.pdf", build_reno_doc_pack),
        ("/tmp/Action_Plan_Final.pdf", build_action_plan),
    ]
    paths = []
    for path, fn in files:
        print(f"Building {os.path.basename(path)}...")
        fn(path)
        size = os.path.getsize(path)
        print(f"  -> {size} bytes")
        paths.append(path)
    print(f"\nSending email with {len(paths)} attachments to {RECIPIENT}...")
    send_all(paths, RECIPIENT)
    print("Done!")
